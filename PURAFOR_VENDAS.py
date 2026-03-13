"""
Relatório de Análise de Vendas por Produto
Empresa: PURAFOR
Lê todos os XMLs de NF-e da pasta e gera um relatório Excel
"""

import os
import json
import hashlib
import threading
import concurrent.futures
import xml.etree.ElementTree as ET
import html as html_mod
from datetime import datetime
import math
import unicodedata
import requests
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint


# ──────────────────────────────────────────────
# CONFIGURAÇÕES
# ──────────────────────────────────────────────
PASTA_XML     = r"Z:\codigos\Fabio\XML"
CATALOGO_XLSX   = r"Z:\codigos\Fabio\Produtos Purafor.xlsx"

# Credenciais: lidas de variável de ambiente (Streamlit Cloud) com fallback local
OMIE_APP_KEY    = os.getenv("OMIE_APP_KEY",    "2786968546362")
OMIE_APP_SECRET = os.getenv("OMIE_APP_SECRET", "0552cb74d4e7dd891a7960a879615385")

# Período de busca das NF-e na API Omie
# Ajuste OMIE_DATA_INI e OMIE_DATA_FIM conforme necessário.
_ano_atual      = datetime.now().year
OMIE_DATA_INI   = f"01/01/{_ano_atual}"          # ex: "01/01/2025"
OMIE_DATA_FIM   = datetime.now().strftime("%d/%m/%Y")  # até hoje

_agora = datetime.now().strftime("%Y%m%d_%H%M%S")
# Usa pasta local se estiver no Windows com Z:\, senão usa pasta relativa ao script
_BASE_SAIDA   = r"Z:\codigos\Fabio" if os.path.isdir(r"Z:\codigos\Fabio") else os.path.dirname(os.path.abspath(__file__))
SAIDA_EXCEL   = os.path.join(_BASE_SAIDA, f"Relatorio_Vendas_PURAFOR_{_agora}.xlsx")
SAIDA_HTML    = os.path.join(_BASE_SAIDA, f"Dashboard_PURAFOR_{_agora}.html")

NS = "http://www.portalfiscal.inf.br/nfe"

# ── Diretório de cache em disco ───────────────────────────────────────────────
# No Streamlit Cloud o repositório é read-only; usa /tmp que é sempre gravável.
# Localmente (Windows) usa a pasta do próprio script.
import tempfile as _tempfile
_CACHE_DIR = os.path.join(
    _tempfile.gettempdir() if os.name != 'nt' else os.path.dirname(os.path.abspath(__file__)),
    '_cache_omie'
)
os.makedirs(_CACHE_DIR, exist_ok=True)

# ── Cache em memória (persiste entre reruns do Streamlit via sys.modules) ──────
# Estrutura: {'data_ini': str, 'data_fim': str, 'records': list, 'saved_at': datetime}
_MEM_VENDAS: dict | None = None
# Estrutura: {'data': dict, 'saved_at': datetime}
_MEM_CATALOGO: dict | None = None

# ── Callback de progresso (injetado pelo Streamlit) ────────────────────────────
# callable(pct: float 0-1, msg: str) ou None
_progresso = None

def _prog(pct: float, msg: str = ""):
    """Emite progresso se o callback estiver configurado."""
    if callable(_progresso):
        try:
            _progresso(min(float(pct), 1.0), msg)
        except Exception:
            pass

# CFOPs que representam vendas reais (ignora remessas/brindes c/ valor simbólico)
CFOP_VENDA = {
    # Venda de produção própria (2023 e anteriores)
    "5101", "6101",
    # Venda de produção própria – não contribuinte / consumidor final
    "6107", "6109",
    # Venda de mercadoria adquirida/recebida de terceiros (2024+)
    "5102", "6102",
    # Venda de mercadoria sujeita a ST / substituição
    "5108", "6108",
    # Venda de mercadoria / operações diversas
    "5110", "6110",
}


# ──────────────────────────────────────────────
# CATÁLOGO DE PRODUTOS (Família + Marca)
# ──────────────────────────────────────────────
def carregar_catalogo(caminho: str) -> pd.DataFrame:
    """
    Lê o Excel de produtos e retorna df com colunas:
    Codigo, Familia, Marca
    Match é feito pelo campo 'Código' (col 2) do catálogo.
    """
    try:
        df = pd.read_excel(caminho, sheet_name=0, usecols=[2, 3, 19], header=0)
        df.columns = ["Codigo", "Familia", "Marca"]
        df["Codigo"] = df["Codigo"].astype(str).str.strip()
        # Mantém apenas linhas com Família ou Marca preenchida
        df = df[df["Familia"].notna() | df["Marca"].notna()].copy()
        df["Familia"] = df["Familia"].fillna("").astype(str).str.strip()
        df["Marca"]   = df["Marca"].fillna("").astype(str).str.strip()
        print(f"  ✔ Catálogo carregado: {len(df)} produtos com Família/Marca")
        return df
    except Exception as e:
        print(f"  [AVISO] Não foi possível carregar o catálogo: {e}")
        return pd.DataFrame(columns=["Codigo", "Familia", "Marca"])


def _norm_cod(s: str) -> str:
    """Normaliza código: remove acentos, maiúsculo, sem espaços extras."""
    return unicodedata.normalize('NFD', str(s)).encode('ascii', 'ignore').decode().upper().strip()


def carregar_catalogo_omie() -> dict:
    """
    Baixa todos os produtos do ERP Omie via API e retorna um dict:
      chave = código normalizado (sem acento, maiúsculo)
      valor = dict completo do produto (todos os campos: familia, marca, ean, ncm, etc.)
    Também indexa variações sem prefixo UN/CX para cobrir diferenças de código entre
    o XML da NF-e e o cadastro do Omie (ex: 'AMORAISO' → 'UNAMORAISO').
    Cache em disco com TTL de 6 horas — na 2ª execução carrega em < 1 s.
    Cache em memória com TTL de 6 horas: instantâneo enquanto o processo estiver vivo.
    """
    global _MEM_CATALOGO
    _TTL_HORAS = 6

    # ── Cache em memória (zero I/O, persiste entre reruns do Streamlit) ──
    if _MEM_CATALOGO is not None:
        _age_h = (datetime.now() - _MEM_CATALOGO['saved_at']).total_seconds() / 3600
        if _age_h < _TTL_HORAS:
            print(f"  ✔ Catálogo Omie (memória {_age_h:.1f}h atrás): "
                  f"{len(_MEM_CATALOGO['data'])} chaves")
            return _MEM_CATALOGO['data']

    _cache_dir  = _CACHE_DIR
    _cache_path = os.path.join(_cache_dir, 'catalogo_omie.json')

    # ── Tenta carregar do cache em disco ──────────────────────────────
    try:
        with open(_cache_path, encoding='utf-8') as _f:
            _raw = json.load(_f)
        _saved_at = datetime.fromisoformat(_raw['saved_at'])
        _age_h = (datetime.now() - _saved_at).total_seconds() / 3600
        if _age_h < _TTL_HORAS:
            omie_map = _raw['data']
            print(f"  ✔ Catálogo Omie (disco {_age_h:.1f}h atrás): {len(omie_map)} chaves")
            _MEM_CATALOGO = {'data': omie_map, 'saved_at': datetime.now()}
            return omie_map
    except Exception:
        pass

    # ── Busca na API ──────────────────────────────────────────────
    URL = 'https://app.omie.com.br/api/v1/geral/produtos/'
    REG_PAG = 50
    try:
        r0 = requests.post(URL, json={
            'call': 'ListarProdutos',
            'app_key': OMIE_APP_KEY,
            'app_secret': OMIE_APP_SECRET,
            'param': [{'pagina': 1, 'registros_por_pagina': REG_PAG, 'filtrar_apenas_omiepdv': 'N'}]
        }, timeout=30)
        total = r0.json().get('total_de_registros', 0)
        tot_pag = math.ceil(total / REG_PAG)
    except Exception as e:
        print(f"  [AVISO] Omie API indisponível: {e}")
        return {}

    omie_map = {}   # _norm_cod(codigo) -> dict completo do produto
    PREFIXOS = ['UN', 'CX']
    for pag in range(1, tot_pag + 1):
        try:
            r = requests.post(URL, json={
                'call': 'ListarProdutos',
                'app_key': OMIE_APP_KEY,
                'app_secret': OMIE_APP_SECRET,
                'param': [{'pagina': pag, 'registros_por_pagina': REG_PAG, 'filtrar_apenas_omiepdv': 'N'}]
            }, timeout=30)
            for p in r.json().get('produto_servico_cadastro', []):
                cod = str(p.get('codigo', '') or '').strip()
                if not cod:
                    continue
                key = _norm_cod(cod)
                omie_map[key] = p
                # Indexa também sem prefixo (UNAMARGO → AMARGO, CXCARVÃO → CARVAO)
                for pref in PREFIXOS:
                    pn = _norm_cod(pref)
                    if key.startswith(pn) and len(key) > len(pn):
                        without = key[len(pn):]
                        if without not in omie_map:
                            omie_map[without] = p
        except Exception as e:
            print(f"  [AVISO] Erro Omie pág {pag}: {e}")

    print(f"  ✔ Omie: {total} produtos baixados ({len(omie_map)} chaves de busca)")

    # ── Salva em memória ────────────────────────────────────────────
    _MEM_CATALOGO = {'data': omie_map, 'saved_at': datetime.now()}

    # ── Persiste cache em disco ─────────────────────────────────────────
    try:
        with open(_cache_path, 'w', encoding='utf-8') as _f:
            json.dump({'saved_at': datetime.now().isoformat(), 'data': omie_map},
                      _f, ensure_ascii=False)
    except Exception:
        pass

    return omie_map


# ──────────────────────────────────────────────
# LEITURA DAS NF-e VIA API OMIE
# ──────────────────────────────────────────────
def _parsear_xml_nfe(xml_str: str) -> list[dict]:
    """
    Recebe o texto XML de uma NF-e e retorna lista de itens de venda.
    Filtra apenas CFOPs em CFOP_VENDA e ignora NF-e canceladas.
    """
    import html as _html
    registros = []
    try:
        root = ET.fromstring(_html.unescape(xml_str))
    except ET.ParseError:
        return registros

    nfe = root.find(f"{{{NS}}}NFe")
    if nfe is None:
        nfe = root
    infnfe = nfe.find(f"{{{NS}}}infNFe")
    if infnfe is None:
        return registros

    ide = infnfe.find(f"{{{NS}}}ide")
    if ide is None:
        return registros

    num_nf = ide.findtext(f"{{{NS}}}nNF", "")
    serie  = ide.findtext(f"{{{NS}}}serie", "")
    dh_emi = ide.findtext(f"{{{NS}}}dhEmi", "")
    try:
        data_emissao = datetime.fromisoformat(dh_emi[:19])
    except Exception:
        data_emissao = None

    dest = infnfe.find(f"{{{NS}}}dest")
    cliente = dest.findtext(f"{{{NS}}}xNome", "") if dest is not None else ""
    uf_dest = ""
    cnpj_dest = ""
    if dest is not None:
        end_dest = dest.find(f"{{{NS}}}enderDest")
        if end_dest is not None:
            uf_dest = end_dest.findtext(f"{{{NS}}}UF", "")
        raw_doc = (dest.findtext(f"{{{NS}}}CNPJ", "")
                   or dest.findtext(f"{{{NS}}}CPF", ""))
        cnpj_dest = raw_doc.replace(".", "").replace("/", "").replace("-", "")

    emit = infnfe.find(f"{{{NS}}}emit")
    emitente = emit.findtext(f"{{{NS}}}xNome", "") if emit is not None else ""

    for det in infnfe.findall(f"{{{NS}}}det"):
        prod = det.find(f"{{{NS}}}prod")
        if prod is None:
            continue
        cfop = prod.findtext(f"{{{NS}}}CFOP", "")
        if cfop not in CFOP_VENDA:
            continue
        cod_prod  = prod.findtext(f"{{{NS}}}cProd", "")
        desc_prod = prod.findtext(f"{{{NS}}}xProd", "")
        unidade   = prod.findtext(f"{{{NS}}}uCom", "")
        try:
            qtd = float(prod.findtext(f"{{{NS}}}qCom", "0"))
        except Exception:
            qtd = 0.0
        try:
            v_unit = float(prod.findtext(f"{{{NS}}}vUnCom", "0"))
        except Exception:
            v_unit = 0.0
        try:
            v_bruto = float(prod.findtext(f"{{{NS}}}vProd", "0"))
        except Exception:
            v_bruto = 0.0
        try:
            v_desc = float(prod.findtext(f"{{{NS}}}vDesc", "0"))
        except Exception:
            v_desc = 0.0
        registros.append({
            "NF":           num_nf,
            "Série":        serie,
            "Data Emissão": data_emissao,
            "Emitente":     emitente,
            "Cliente":      cliente,
            "UF Dest.":     uf_dest,
            "CNPJ_Dest":    cnpj_dest,
            "CFOP":         cfop,
            "Cód. Produto": cod_prod,
            "Produto":      desc_prod,
            "Família":      "",
            "Marca":        "",
            "Unidade":      unidade,
            "Qtd":          qtd,
            "Vlr Unitário": v_unit,
            "Vlr Bruto":    v_bruto,
            "Desconto":     v_desc,
            "Vlr Líquido":  v_bruto - v_desc,
        })
    return registros


def ler_xmls_omie_api(data_ini: str, data_fim: str) -> list[dict]:
    """
    Baixa todas as NF-e modelo 55 autorizadas do Omie no período informado.
    Substitui a leitura da pasta local de XMLs.
    - data_ini / data_fim: formato "DD/MM/AAAA"
    - NF-e canceladas (cStatus='40') são ignoradas.
    - Filtra itens com CFOP em CFOP_VENDA.
    - Usa 500 registros/página + busca paralela (5 workers) para máxima velocidade.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading

    URL = 'https://app.omie.com.br/api/v1/contador/xml/'
    REG_PAG = 500  # máximo Omie — reduz 41 páginas → ~5 páginas

    def _fetch_pag(pag_num: int) -> tuple:
        """Busca uma página e retorna (numero_pagina, resposta)."""
        resp = requests.post(URL, json={
            'call': 'ListarDocumentos',
            'app_key': OMIE_APP_KEY,
            'app_secret': OMIE_APP_SECRET,
            'param': [{
                'cModelo': '55',
                'dEmiInicial': data_ini,
                'dEmiFinal': data_fim,
                'nPagina': pag_num,
                'nRegPorPagina': REG_PAG,
            }]
        }, timeout=120).json()
        return pag_num, resp

    def _processar_resp(resp: dict) -> list:
        """Extrai itens de uma resposta, ignorando NFs canceladas."""
        itens_pag = []
        for doc in resp.get('documentosEncontrados', []):
            if doc.get('cStatus') == '40':
                continue
            xml_str = doc.get('cXml', '')
            n_id_pedido = doc.get('nIdPedido', 0)
            n_chave = doc.get('nChave', '')
            if xml_str:
                itens = _parsear_xml_nfe(xml_str)
                for item in itens:
                    item['nIdPedido'] = n_id_pedido
                    item['nChave'] = n_chave
                itens_pag.extend(itens)
        return itens_pag

    # ── Página 1: descobre total de páginas ──────────────────────────────────
    try:
        _, resp0 = _fetch_pag(1)
        if 'faultstring' in resp0:
            print(f"  [ERRO] Omie API: {resp0['faultstring']}")
            return []
        total   = resp0.get('nTotRegistros', 0)
        tot_pag = resp0.get('nTotPaginas', math.ceil(total / REG_PAG))
        print(f"  Omie: {total} NF-e no período ({tot_pag} páginas de {REG_PAG})")
    except Exception as e:
        print(f"  [ERRO] Omie API indisponível: {e}")
        return []

    _prog(0.08, f"Vendas: página 1/{tot_pag}...")

    # ── Páginas 2..N em paralelo (5 workers simultâneos) ─────────────────────
    _lock = threading.Lock()
    resultados_por_pag = {1: _processar_resp(resp0)}
    _concluidas = [1]  # contador protegido por _lock

    if tot_pag > 1:
        with ThreadPoolExecutor(max_workers=5) as ex:
            futures = {ex.submit(_fetch_pag, p): p for p in range(2, tot_pag + 1)}
            for fut in as_completed(futures):
                try:
                    pag_num, resp = fut.result()
                    itens_pag = _processar_resp(resp)
                    with _lock:
                        resultados_por_pag[pag_num] = itens_pag
                        _concluidas.append(pag_num)
                        n_done = len(_concluidas)
                    _prog(0.05 + (n_done / max(tot_pag, 1)) * 0.33,
                          f"Vendas: {n_done}/{tot_pag} páginas concluídas...")
                    print(f"    pág {pag_num}/{tot_pag} OK ({len(itens_pag)} itens)")
                except Exception as e:
                    pag_num = futures[fut]
                    print(f"  [AVISO] Erro pág {pag_num}: {e}")
                    with _lock:
                        resultados_por_pag[pag_num] = []

    # ── Ordena por página e junta tudo ───────────────────────────────────────
    registros = []
    for p in sorted(resultados_por_pag):
        registros.extend(resultados_por_pag[p])

    print(f"  ✔ {len(registros)} itens de {tot_pag} páginas baixadas em paralelo")
    return registros


# ──────────────────────────────────────────────
# CACHE INCREMENTAL DE VENDAS
# ──────────────────────────────────────────────
_DIAS_INCREMENTAL = 45  # ao usar cache, rebusca sempre os últimos N dias

def _ler_vendas_com_cache(data_ini: str, data_fim: str) -> list[dict]:
    """
    Wrapper incremental sobre ler_xmls_omie_api:
    - 1ª execução: baixa tudo e salva em _cache_omie/vendas_full.json
    - Execuções seguintes: carrega cache + rebusca apenas últimos
      _DIAS_INCREMENTAL dias para pegar NFs novas/alteradas.
    - Cache em memória: se mesmo período e < 30 min, retorna instantâneo.
    - Filtra resultado pelo período solicitado.
    """
    global _MEM_VENDAS
    from datetime import timedelta

    _DT_FMT = '%d/%m/%Y'
    _ISO_FMT = '%Y-%m-%dT%H:%M:%S'
    _MEM_TTL_MIN = 30   # minutos antes de rebuscar incremento

    # ── Cache em memória: período coberto pelo conjunto completo e recente? ──
    # Guarda TODOS os registros baixados; filtra aqui sem novos downloads.
    if _MEM_VENDAS is not None:
        _age_min = (datetime.now() - _MEM_VENDAS['saved_at']).total_seconds() / 60
        if _age_min < _MEM_TTL_MIN:
            _mem_ini = _MEM_VENDAS.get('earliest')
            _mem_fim = _MEM_VENDAS.get('latest')
            d_ini_req = datetime.strptime(data_ini, _DT_FMT)
            d_fim_req = datetime.strptime(data_fim, _DT_FMT)
            if (
                _mem_ini is not None and _mem_fim is not None
                and d_ini_req >= _mem_ini
                and d_fim_req <= _mem_fim
            ):
                resultado = [
                    r for r in _MEM_VENDAS['records_all']
                    if r.get('Data Emissão') is not None
                    and d_ini_req <= r['Data Emissão'] <= d_fim_req
                ]
                print(f"  ✔ Vendas (memória {_age_min:.1f} min, filtrado): "
                      f"{len(resultado)} registros")
                _prog(0.38, "Vendas carregadas do cache em memória")
                return resultado

    def _to_str(r: dict) -> dict:
        rc = dict(r)
        if isinstance(rc.get('Data Emissão'), datetime):
            rc['Data Emissão'] = rc['Data Emissão'].strftime(_ISO_FMT)
        return rc

    def _from_str(r: dict) -> dict:
        rc = dict(r)
        v = rc.get('Data Emissão')
        if isinstance(v, str):
            try:
                rc['Data Emissão'] = datetime.fromisoformat(v[:19])
            except Exception:
                pass
        return rc

    d_ini = datetime.strptime(data_ini, _DT_FMT)
    d_fim = datetime.strptime(data_fim, _DT_FMT)

    _cache_dir  = _CACHE_DIR
    _cache_path = os.path.join(_cache_dir, 'vendas_v3.json')  # v3: re-fetch com nChave

    all_cached: list[dict] = []
    cache_earliest: datetime | None = None
    if os.path.exists(_cache_path):
        try:
            with open(_cache_path, encoding='utf-8') as f:
                raw = json.load(f)
            all_cached = [_from_str(r) for r in raw.get('records', [])]
            meta = raw.get('meta', {})
            updated = meta.get('updated', '?')
            _e = meta.get('earliest')
            if _e:
                try:
                    cache_earliest = datetime.fromisoformat(_e[:19])
                except Exception:
                    pass
            print(f"  Cache vendas: {len(all_cached)} registros "
                  f"(salvo {updated[:10]}, "
                  f"desde {cache_earliest.strftime(_DT_FMT) if cache_earliest else '?'})")
        except Exception as e:
            print(f"  [AVISO] Cache de vendas corrompido, ignorando: {e}")
            all_cached = []
            cache_earliest = None

    # ── Estratégia: incremental ou full fetch ──────────────────────
    # Força full fetch APENAS se cache vazio OU se conhecemos a data mais antiga
    # do cache E o período solicitado começa antes dela.
    # Se cache_earliest é None mas há registros, assume incremental (cache sem metadado antigo).
    _need_full = (
        not all_cached
        or (cache_earliest is not None and d_ini < cache_earliest)
    )
    if _need_full:
        if all_cached:
            _e_str = cache_earliest.strftime(_DT_FMT) if cache_earliest else '?'
            print(f"  Período solicitado ({data_ini}) antes do cache "
                  f"({_e_str}) — full fetch")
        else:
            print(f"  Cache vazio — buscando período completo: {data_ini} → {data_fim}")
        _prog(0.05, f"Buscando vendas: {data_ini} → {data_fim}...")
        all_records = ler_xmls_omie_api(data_ini, data_fim)
    else:
        incr_start = d_fim - timedelta(days=_DIAS_INCREMENTAL)
        incr_ini_str = incr_start.strftime(_DT_FMT)
        print(f"  Incremento: buscando {incr_ini_str} → {data_fim} (últimos {_DIAS_INCREMENTAL} dias)")
        _prog(0.05, f"Buscando incremento: {incr_ini_str} → {data_fim}...")
        novos = ler_xmls_omie_api(incr_ini_str, data_fim)
        # Mantém do cache apenas o que está ANTES da janela incremental
        sobreviventes = [
            r for r in all_cached
            if r.get('Data Emissão') is not None
            and r['Data Emissão'] < incr_start
        ]
        all_records = sobreviventes + novos

    # ── Deduplica por NF+Série+Produto+Data ───────────────────────
    seen: set = set()
    dedup: list[dict] = []
    for r in all_records:
        k = (
            r.get('NF', ''),
            r.get('Série', ''),
            r.get('Cód. Produto', ''),
            str(r.get('Data Emissão', ''))[:10],
        )
        if k not in seen:
            seen.add(k)
            dedup.append(r)

    # ── Persiste cache atualizado ─────────────────────────────────
    try:
        _datas = [r['Data Emissão'] for r in dedup if isinstance(r.get('Data Emissão'), datetime)]
        _earliest_iso = min(_datas).isoformat() if _datas else None
        with open(_cache_path, 'w', encoding='utf-8') as f:
            json.dump(
                {
                    'meta': {
                        'updated':  datetime.now().isoformat(),
                        'total':    len(dedup),
                        'earliest': _earliest_iso,
                    },
                    'records': [_to_str(r) for r in dedup],
                },
                f, ensure_ascii=False, default=str,
            )
        print(f"  Cache vendas atualizado: {len(dedup)} registros "
              f"(desde {_earliest_iso[:10] if _earliest_iso else '?'})")
    except Exception as e:
        print(f"  [AVISO] Não foi possível salvar cache de vendas: {e}")

    # ── Salva em memória o conjunto COMPLETO ─────────────────────
    _datas_mem = [r['Data Emissão'] for r in dedup if isinstance(r.get('Data Emissão'), datetime)]
    _MEM_VENDAS = {
        'records_all': dedup,
        'earliest':    min(_datas_mem) if _datas_mem else None,
        'latest':      max(_datas_mem) if _datas_mem else None,
        'saved_at':    datetime.now(),
    }

    # ── Filtra pelo período solicitado ────────────────────────────
    resultado = [
        r for r in dedup
        if r.get('Data Emissão') is not None
        and d_ini <= r['Data Emissão'] <= d_fim
    ]
    return resultado


# ──────────────────────────────────────────────
# LEITURA DOS XMLs (pasta local — mantido como fallback)
# ──────────────────────────────────────────────
def ler_xmls(pasta: str) -> list[dict]:
    registros = []

    for nome_arquivo in sorted(os.listdir(pasta)):
        if not nome_arquivo.lower().endswith(".xml"):
            continue
        # ignora eventos (cancelamento etc.)
        if "procEventoNFe" in nome_arquivo:
            continue

        caminho = os.path.join(pasta, nome_arquivo)
        try:
            tree = ET.parse(caminho)
            root = tree.getroot()

            # Tag raiz pode ser nfeProc ou NFe diretamente
            nfe = root.find(f"{{{NS}}}NFe")
            if nfe is None:
                nfe = root  # arquivo já é NFe

            infnfe = nfe.find(f"{{{NS}}}infNFe")
            if infnfe is None:
                continue

            # ── Cabeçalho ──────────────────────────────
            ide = infnfe.find(f"{{{NS}}}ide")
            num_nf = ide.findtext(f"{{{NS}}}nNF", "")
            serie = ide.findtext(f"{{{NS}}}serie", "")
            dh_emi = ide.findtext(f"{{{NS}}}dhEmi", "")
            # parse da data
            try:
                data_emissao = datetime.fromisoformat(dh_emi[:19])
            except Exception:
                data_emissao = None

            # ── Destinatário ────────────────────────────
            dest = infnfe.find(f"{{{NS}}}dest")
            cliente = dest.findtext(f"{{{NS}}}xNome", "") if dest is not None else ""
            uf_dest = ""
            if dest is not None:
                end_dest = dest.find(f"{{{NS}}}enderDest")
                if end_dest is not None:
                    uf_dest = end_dest.findtext(f"{{{NS}}}UF", "")

            # ── Emitente ────────────────────────────────
            emit = infnfe.find(f"{{{NS}}}emit")
            emitente = emit.findtext(f"{{{NS}}}xNome", "") if emit is not None else ""

            # ── Itens ───────────────────────────────────
            for det in infnfe.findall(f"{{{NS}}}det"):
                prod = det.find(f"{{{NS}}}prod")
                if prod is None:
                    continue

                cfop = prod.findtext(f"{{{NS}}}CFOP", "")
                # Filtra apenas CFOPs de venda
                if cfop not in CFOP_VENDA:
                    continue

                cod_prod = prod.findtext(f"{{{NS}}}cProd", "")
                desc_prod = prod.findtext(f"{{{NS}}}xProd", "")
                unidade = prod.findtext(f"{{{NS}}}uCom", "")
                try:
                    qtd = float(prod.findtext(f"{{{NS}}}qCom", "0"))
                except Exception:
                    qtd = 0.0
                try:
                    v_unit = float(prod.findtext(f"{{{NS}}}vUnCom", "0"))
                except Exception:
                    v_unit = 0.0
                try:
                    v_bruto = float(prod.findtext(f"{{{NS}}}vProd", "0"))
                except Exception:
                    v_bruto = 0.0
                try:
                    v_desc = float(prod.findtext(f"{{{NS}}}vDesc", "0"))
                except Exception:
                    v_desc = 0.0

                v_liquido = v_bruto - v_desc

                registros.append({
                    "NF": num_nf,
                    "Série": serie,
                    "Data Emissão": data_emissao,
                    "Emitente": emitente,
                    "Cliente": cliente,
                    "UF Dest.": uf_dest,
                    "CFOP": cfop,
                    "Cód. Produto": cod_prod,
                    "Produto": desc_prod,
                    "Família": "",
                    "Marca": "",
                    "Unidade": unidade,
                    "Qtd": qtd,
                    "Vlr Unitário": v_unit,
                    "Vlr Bruto": v_bruto,
                    "Desconto": v_desc,
                    "Vlr Líquido": v_liquido,
                })

        except Exception as e:
            print(f"  [AVISO] Erro ao ler {nome_arquivo}: {e}")

    return registros




# ──────────────────────────────────────────────
# ESTILOS
# ──────────────────────────────────────────────
COR_HEADER = "1F4E79"      # azul escuro
COR_HEADER_FONT = "FFFFFF"
COR_TOTAL = "D6E4F0"       # azul claro
COR_ZEBRA = "EBF4FB"
COR_SUBTOTAL = "BDD7EE"


def estilo_header(ws, row, col_ini, col_fim, texto, merge=True):
    if merge:
        ws.merge_cells(start_row=row, start_column=col_ini,
                       end_row=row, end_column=col_fim)
    cell = ws.cell(row=row, column=col_ini, value=texto)
    cell.font = Font(bold=True, color=COR_HEADER_FONT, size=12)
    cell.fill = PatternFill("solid", fgColor=COR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def aplicar_borda(ws, min_row, max_row, min_col, max_col):
    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = borda


def largura_auto(ws, col_larguras: dict):
    for col_letra, largura in col_larguras.items():
        ws.column_dimensions[col_letra].width = largura


# ──────────────────────────────────────────────
# SHEET 1 — Detalhe de Vendas
# ──────────────────────────────────────────────
def sheet_detalhe(wb, df: pd.DataFrame):
    ws = wb.create_sheet("Detalhe de Vendas")

    colunas = [
        "NF", "Data Emissão", "Cliente", "UF Dest.",
        "Cód. Produto", "Produto", "Família", "Marca", "Unidade",
        "Qtd", "Vlr Unitário", "Vlr Bruto", "Desconto", "Vlr Líquido"
    ]

    # Título
    estilo_header(ws, 1, 1, len(colunas), "DETALHE DE VENDAS — PURAFOR")
    ws.row_dimensions[1].height = 25

    # Cabeçalho das colunas
    h_fill = PatternFill("solid", fgColor=COR_HEADER)
    h_font = Font(bold=True, color=COR_HEADER_FONT, size=10)
    for ci, col in enumerate(colunas, 1):
        cell = ws.cell(row=2, column=ci, value=col)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    # Dados
    z_fill = PatternFill("solid", fgColor=COR_ZEBRA)
    for ri, row in enumerate(df[colunas].itertuples(index=False), start=3):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical="center")
            if ri % 2 == 0:
                cell.fill = z_fill
            # Formatos
            if ci == 2 and val:          # Data
                cell.number_format = "DD/MM/YYYY"
            elif ci in (8,):             # Qtd
                cell.number_format = "#,##0.00"
            elif ci in (9, 10, 11, 12):  # Valores monetários
                cell.number_format = 'R$ #,##0.00'

    # Linha de total  (colunas: A=NF B=Data C=Cliente D=UF E=Cod F=Prod G=Familia H=Marca I=Unid J=Qtd K=VUn L=VBruto M=Desc N=VLiq)
    total_row = len(df) + 3
    ws.cell(row=total_row, column=9, value="TOTAL").font = Font(bold=True)
    for ci in [10, 12, 13, 14]:   # Qtd, VBruto, Desc, VLiq
        cell = ws.cell(row=total_row, column=ci,
                       value=f"=SUM({get_column_letter(ci)}3:{get_column_letter(ci)}{total_row-1})")
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=COR_TOTAL)
        cell.number_format = "#,##0.00" if ci == 10 else 'R$ #,##0.00'

    aplicar_borda(ws, 2, total_row, 1, len(colunas))

    # Larguras
    larguras = {
        "A": 8, "B": 14, "C": 35, "D": 8,
        "E": 16, "F": 42, "G": 22, "H": 18, "I": 9,
        "J": 10, "K": 14, "L": 14, "M": 14, "N": 14
    }
    largura_auto(ws, larguras)
    ws.freeze_panes = "A3"


# ──────────────────────────────────────────────
# SHEET 2 — Resumo por Produto
# ──────────────────────────────────────────────
def sheet_por_produto(wb, df: pd.DataFrame):
    ws = wb.create_sheet("Vendas por Produto")

    grp = (
        df.groupby(["Cód. Produto", "Produto"])
        .agg(
            Qtd_Total=("Qtd", "sum"),
            Vlr_Bruto=("Vlr Bruto", "sum"),
            Desconto=("Desconto", "sum"),
            Vlr_Liquido=("Vlr Líquido", "sum"),
            Num_NFs=("NF", "nunique"),
        )
        .reset_index()
        .sort_values("Vlr_Liquido", ascending=False)
    )
    grp["Part_%"] = grp["Vlr_Liquido"] / grp["Vlr_Liquido"].sum() * 100
    grp["Part_% Acum."] = grp["Part_%"].cumsum()

    cols = [
        "Cód. Produto", "Produto", "Qtd_Total",
        "Vlr_Bruto", "Desconto", "Vlr_Liquido",
        "Part_%", "Part_% Acum.", "Num_NFs"
    ]
    headers = [
        "Cód. Produto", "Produto", "Qtd Total",
        "Vlr Bruto", "Desconto", "Vlr Líquido",
        "Part. %", "Part. % Acum.", "Nº NFs"
    ]

    estilo_header(ws, 1, 1, len(cols), "RANKING DE VENDAS POR PRODUTO — PURAFOR")
    ws.row_dimensions[1].height = 25

    h_fill = PatternFill("solid", fgColor=COR_HEADER)
    h_font = Font(bold=True, color=COR_HEADER_FONT, size=10)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    z_fill = PatternFill("solid", fgColor=COR_ZEBRA)
    for ri, row in enumerate(grp[cols].itertuples(index=False), start=3):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=round(val, 4) if isinstance(val, float) else val)
            cell.alignment = Alignment(vertical="center")
            if ri % 2 == 0:
                cell.fill = z_fill
            if ci == 3:
                cell.number_format = "#,##0.00"
            elif ci in (4, 5, 6):
                cell.number_format = 'R$ #,##0.00'
            elif ci in (7, 8):
                cell.number_format = "0.00%"
                cell.value = (val / 100) if isinstance(val, float) else val

    # Total
    total_row = len(grp) + 3
    ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
    for ci, col_letra in [(3, "C"), (4, "D"), (5, "E"), (6, "F")]:
        cell = ws.cell(row=total_row, column=ci,
                       value=f"=SUM({col_letra}3:{col_letra}{total_row-1})")
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=COR_TOTAL)
        fmt = "#,##0.00" if ci == 3 else 'R$ #,##0.00'
        cell.number_format = fmt

    aplicar_borda(ws, 2, total_row, 1, len(cols))

    larguras = {
        "A": 16, "B": 44, "C": 12,
        "D": 14, "E": 14, "F": 14,
        "G": 10, "H": 13, "I": 8,
    }
    largura_auto(ws, larguras)
    ws.freeze_panes = "A3"

    return grp  # retorna para o gráfico


# ──────────────────────────────────────────────
# SHEET 3 — Resumo por Cliente
# ──────────────────────────────────────────────
def sheet_por_cliente(wb, df: pd.DataFrame):
    ws = wb.create_sheet("Vendas por Cliente")

    grp = (
        df.groupby(["Cliente", "UF Dest."])
        .agg(
            Num_NFs=("NF", "nunique"),
            Vlr_Bruto=("Vlr Bruto", "sum"),
            Desconto=("Desconto", "sum"),
            Vlr_Liquido=("Vlr Líquido", "sum"),
        )
        .reset_index()
        .sort_values("Vlr_Liquido", ascending=False)
    )
    grp["Part_%"] = grp["Vlr_Liquido"] / grp["Vlr_Liquido"].sum() * 100

    cols = ["Cliente", "UF Dest.", "Num_NFs", "Vlr_Bruto", "Desconto", "Vlr_Liquido", "Part_%"]
    headers = ["Cliente", "UF", "Nº NFs", "Vlr Bruto", "Desconto", "Vlr Líquido", "Part. %"]

    estilo_header(ws, 1, 1, len(cols), "RANKING DE VENDAS POR CLIENTE — PURAFOR")
    ws.row_dimensions[1].height = 25

    h_fill = PatternFill("solid", fgColor=COR_HEADER)
    h_font = Font(bold=True, color=COR_HEADER_FONT, size=10)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    z_fill = PatternFill("solid", fgColor=COR_ZEBRA)
    for ri, row in enumerate(grp[cols].itertuples(index=False), start=3):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=round(val, 4) if isinstance(val, float) else val)
            cell.alignment = Alignment(vertical="center")
            if ri % 2 == 0:
                cell.fill = z_fill
            if ci in (4, 5, 6):
                cell.number_format = 'R$ #,##0.00'
            elif ci == 7:
                cell.number_format = "0.00%"
                cell.value = val / 100

    total_row = len(grp) + 3
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for ci, col_letra in [(4, "D"), (5, "E"), (6, "F")]:
        cell = ws.cell(row=total_row, column=ci,
                       value=f"=SUM({col_letra}3:{col_letra}{total_row-1})")
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=COR_TOTAL)
        cell.number_format = 'R$ #,##0.00'

    aplicar_borda(ws, 2, total_row, 1, len(cols))
    largura_auto(ws, {"A": 42, "B": 6, "C": 8, "D": 14, "E": 14, "F": 14, "G": 10})
    ws.freeze_panes = "A3"


# ──────────────────────────────────────────────
# SHEET 4 — Resumo por Data
# ──────────────────────────────────────────────
def sheet_por_data(wb, df: pd.DataFrame):
    ws = wb.create_sheet("Vendas por Data")

    df2 = df.copy()
    df2["Data"] = pd.to_datetime(df2["Data Emissão"]).dt.date

    grp = (
        df2.groupby("Data")
        .agg(
            Num_NFs=("NF", "nunique"),
            Vlr_Bruto=("Vlr Bruto", "sum"),
            Desconto=("Desconto", "sum"),
            Vlr_Liquido=("Vlr Líquido", "sum"),
        )
        .reset_index()
        .sort_values("Data")
    )

    cols = ["Data", "Num_NFs", "Vlr_Bruto", "Desconto", "Vlr_Liquido"]
    headers = ["Data", "Nº NFs", "Vlr Bruto", "Desconto", "Vlr Líquido"]

    estilo_header(ws, 1, 1, len(cols), "VENDAS POR DATA — PURAFOR")
    ws.row_dimensions[1].height = 25

    h_fill = PatternFill("solid", fgColor=COR_HEADER)
    h_font = Font(bold=True, color=COR_HEADER_FONT, size=10)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28

    z_fill = PatternFill("solid", fgColor=COR_ZEBRA)
    for ri, row in enumerate(grp[cols].itertuples(index=False), start=3):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical="center")
            if ri % 2 == 0:
                cell.fill = z_fill
            if ci == 1:
                cell.number_format = "DD/MM/YYYY"
            elif ci in (3, 4, 5):
                cell.number_format = 'R$ #,##0.00'

    total_row = len(grp) + 3
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for ci, col_letra in [(3, "C"), (4, "D"), (5, "E")]:
        cell = ws.cell(row=total_row, column=ci,
                       value=f"=SUM({col_letra}3:{col_letra}{total_row-1})")
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=COR_TOTAL)
        cell.number_format = 'R$ #,##0.00'

    aplicar_borda(ws, 2, total_row, 1, len(cols))
    largura_auto(ws, {"A": 14, "B": 8, "C": 14, "D": 14, "E": 14})
    ws.freeze_panes = "A3"


# ──────────────────────────────────────────────
# SHEET — Vendas por Família
# ──────────────────────────────────────────────
def sheet_por_familia(wb, df: pd.DataFrame):
    ws = wb.create_sheet("Vendas por Família")

    grp = (
        df.groupby("Família")
        .agg(
            Num_NFs=("NF", "nunique"),
            Qtd_Total=("Qtd", "sum"),
            Vlr_Bruto=("Vlr Bruto", "sum"),
            Desconto=("Desconto", "sum"),
            Vlr_Liquido=("Vlr Líquido", "sum"),
        )
        .reset_index()
        .sort_values("Vlr_Liquido", ascending=False)
    )
    grp["Part_%"] = grp["Vlr_Liquido"] / grp["Vlr_Liquido"].sum() * 100

    cols    = ["Família", "Num_NFs", "Qtd_Total", "Vlr_Bruto", "Desconto", "Vlr_Liquido", "Part_%"]
    headers = ["Família", "Nº NFs", "Qtd Total", "Vlr Bruto", "Desconto", "Vlr Líquido", "Part. %"]

    estilo_header(ws, 1, 1, len(cols), "VENDAS POR FAMÍLIA DE PRODUTO — PURAFOR")
    ws.row_dimensions[1].height = 25

    h_fill = PatternFill("solid", fgColor=COR_HEADER)
    h_font = Font(bold=True, color=COR_HEADER_FONT, size=10)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.fill = h_fill; cell.font = h_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    z_fill = PatternFill("solid", fgColor=COR_ZEBRA)
    for ri, row in enumerate(grp[cols].itertuples(index=False), start=3):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=round(val, 4) if isinstance(val, float) else val)
            cell.alignment = Alignment(vertical="center")
            if ri % 2 == 0: cell.fill = z_fill
            if ci == 3:   cell.number_format = "#,##0.00"
            elif ci in (4, 5, 6): cell.number_format = 'R$ #,##0.00'
            elif ci == 7:
                cell.number_format = "0.00%"
                cell.value = val / 100

    total_row = len(grp) + 3
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for ci, cl in [(3, "C"), (4, "D"), (5, "E"), (6, "F")]:
        cell = ws.cell(row=total_row, column=ci,
                       value=f"=SUM({cl}3:{cl}{total_row-1})")
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=COR_TOTAL)
        cell.number_format = "#,##0.00" if ci == 3 else 'R$ #,##0.00'

    aplicar_borda(ws, 2, total_row, 1, len(cols))
    largura_auto(ws, {"A": 28, "B": 8, "C": 12, "D": 14, "E": 14, "F": 14, "G": 10})
    ws.freeze_panes = "A3"


# ──────────────────────────────────────────────
# SHEET — Vendas por Marca
# ──────────────────────────────────────────────
def sheet_por_marca(wb, df: pd.DataFrame):
    ws = wb.create_sheet("Vendas por Marca")

    grp = (
        df.groupby("Marca")
        .agg(
            Num_NFs=("NF", "nunique"),
            Qtd_Total=("Qtd", "sum"),
            Vlr_Bruto=("Vlr Bruto", "sum"),
            Desconto=("Desconto", "sum"),
            Vlr_Liquido=("Vlr Líquido", "sum"),
        )
        .reset_index()
        .sort_values("Vlr_Liquido", ascending=False)
    )
    grp["Part_%"] = grp["Vlr_Liquido"] / grp["Vlr_Liquido"].sum() * 100

    cols    = ["Marca", "Num_NFs", "Qtd_Total", "Vlr_Bruto", "Desconto", "Vlr_Liquido", "Part_%"]
    headers = ["Marca", "Nº NFs", "Qtd Total", "Vlr Bruto", "Desconto", "Vlr Líquido", "Part. %"]

    estilo_header(ws, 1, 1, len(cols), "VENDAS POR MARCA — PURAFOR")
    ws.row_dimensions[1].height = 25

    h_fill = PatternFill("solid", fgColor=COR_HEADER)
    h_font = Font(bold=True, color=COR_HEADER_FONT, size=10)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.fill = h_fill; cell.font = h_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    z_fill = PatternFill("solid", fgColor=COR_ZEBRA)
    for ri, row in enumerate(grp[cols].itertuples(index=False), start=3):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=round(val, 4) if isinstance(val, float) else val)
            cell.alignment = Alignment(vertical="center")
            if ri % 2 == 0: cell.fill = z_fill
            if ci == 3:   cell.number_format = "#,##0.00"
            elif ci in (4, 5, 6): cell.number_format = 'R$ #,##0.00'
            elif ci == 7:
                cell.number_format = "0.00%"
                cell.value = val / 100

    total_row = len(grp) + 3
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for ci, cl in [(3, "C"), (4, "D"), (5, "E"), (6, "F")]:
        cell = ws.cell(row=total_row, column=ci,
                       value=f"=SUM({cl}3:{cl}{total_row-1})")
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=COR_TOTAL)
        cell.number_format = "#,##0.00" if ci == 3 else 'R$ #,##0.00'

    aplicar_borda(ws, 2, total_row, 1, len(cols))
    largura_auto(ws, {"A": 24, "B": 8, "C": 12, "D": 14, "E": 14, "F": 14, "G": 10})
    ws.freeze_panes = "A3"


# ──────────────────────────────────────────────
# SHEET — De-Para (produtos sem mapeamento)
# ──────────────────────────────────────────────
def sheet_depara(wb, df: pd.DataFrame):
    """Lista produtos sem Família/Marca para o usuário preencher manualmente."""
    ws = wb.create_sheet("De-Para (Preencher)")

    sem_map = (
        df[df["Família"] == "SEM CADASTRO"][["Cód. Produto", "Produto", "Família", "Marca", "Vlr Líquido"]]
        .groupby(["Cód. Produto", "Produto"])
        .agg(Vlr_Liquido=("Vlr Líquido", "sum"))
        .reset_index()
        .sort_values("Vlr_Liquido", ascending=False)
    )

    headers = ["Cód. Produto", "Produto (XML)", "Vlr Líquido", "FAMÍLIA (preencher)", "MARCA (preencher)"]
    aviso = ("⚠ Estes produtos NÃO foram encontrados no catálogo. "
             "Preencha as colunas D e E e execute novamente para ter as análises por Família e Marca completas.")

    ws.merge_cells("A1:E1")
    cell = ws.cell(row=1, column=1, value=aviso)
    cell.font = Font(bold=True, color="9C0006", size=10)
    cell.fill = PatternFill("solid", fgColor="FFC7CE")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 36

    h_fill = PatternFill("solid", fgColor=COR_HEADER)
    h_font = Font(bold=True, color="FFFFFF", size=10)
    preencher_fill = PatternFill("solid", fgColor="FFEB9C")   # amarelo — campos a preencher
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.fill = h_fill; cell.font = h_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    z_fill = PatternFill("solid", fgColor=COR_ZEBRA)
    for ri, row in enumerate(sem_map.itertuples(index=False), start=3):
        for ci, val in enumerate([row[0], row[1], row[2], "", ""], 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical="center")
            if ci == 3: cell.number_format = 'R$ #,##0.00'
            if ci in (4, 5): cell.fill = preencher_fill   # campo a preencher
            elif ri % 2 == 0: cell.fill = z_fill

    aplicar_borda(ws, 2, max(len(sem_map) + 2, 3), 1, 5)
    largura_auto(ws, {"A": 18, "B": 50, "C": 14, "D": 26, "E": 22})
    ws.freeze_panes = "A3"

    print(f"  ⚠ {len(sem_map)} produtos SEM Família/Marca — veja a aba 'De-Para (Preencher)'")


# ──────────────────────────────────────────────
# SHEET 5 — Dashboard / Painel
# ──────────────────────────────────────────────
def sheet_dashboard(wb, df: pd.DataFrame, grp_produto: pd.DataFrame):
    ws = wb.create_sheet("Dashboard")

    total_nfs = df["NF"].nunique()
    total_bruto = df["Vlr Bruto"].sum()
    total_desc = df["Desconto"].sum()
    total_liq = df["Vlr Líquido"].sum()
    total_clientes = df["Cliente"].nunique()
    total_produtos = df["Cód. Produto"].nunique()
    perc_desc = (total_desc / total_bruto * 100) if total_bruto else 0

    estilo_header(ws, 1, 1, 8, "DASHBOARD — ANÁLISE DE VENDAS PURAFOR")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:H1")

    # KPIs
    kpis = [
        ("Total de NFs", total_nfs, ""),
        ("Clientes Atendidos", total_clientes, ""),
        ("Produtos Vendidos", total_produtos, ""),
        ("Faturamento Bruto", total_bruto, "R$"),
        ("Total de Descontos", total_desc, "R$"),
        ("Faturamento Líquido", total_liq, "R$"),
        ("% de Desconto Médio", perc_desc / 100, "%"),
    ]

    kpi_fill = PatternFill("solid", fgColor="D6E4F0")
    kpi_fill2 = PatternFill("solid", fgColor="1F4E79")
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 28

    for i, (label, valor, prefixo) in enumerate(kpis, 1):
        col = i
        c_label = ws.cell(row=3, column=col, value=label)
        c_label.fill = kpi_fill2
        c_label.font = Font(bold=True, color="FFFFFF", size=9)
        c_label.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        c_val = ws.cell(row=4, column=col, value=valor)
        c_val.fill = kpi_fill
        c_val.font = Font(bold=True, size=11)
        c_val.alignment = Alignment(horizontal="center", vertical="center")
        if prefixo == "R$":
            c_val.number_format = 'R$ #,##0.00'
        elif prefixo == "%":
            c_val.number_format = "0.00%"

    # Top 10 produtos — tabela
    top10 = grp_produto.head(10)[["Produto", "Vlr_Liquido"]].reset_index(drop=True)

    ws.cell(row=6, column=1, value="TOP 10 PRODUTOS POR FATURAMENTO LÍQUIDO").font = Font(bold=True, size=11, color=COR_HEADER)
    headers_top = ["Produto", "Vlr Líquido"]
    h_fill = PatternFill("solid", fgColor=COR_HEADER)
    for ci, h in enumerate(headers_top, 1):
        c = ws.cell(row=7, column=ci, value=h)
        c.fill = h_fill
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")

    z_fill = PatternFill("solid", fgColor=COR_ZEBRA)
    for ri, (_, row) in enumerate(top10.iterrows(), start=8):
        c1 = ws.cell(row=ri, column=1, value=row["Produto"])
        c2 = ws.cell(row=ri, column=2, value=row["Vlr_Liquido"])
        c2.number_format = 'R$ #,##0.00'
        if ri % 2 == 0:
            c1.fill = z_fill
            c2.fill = z_fill

    aplicar_borda(ws, 7, 17, 1, 2)

    # Gráfico de barras Top 10
    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "clustered"
    chart.title = "Top 10 Produtos — Faturamento Líquido"
    chart.y_axis.title = "Produto"
    chart.x_axis.title = "R$"
    chart.height = 12
    chart.width = 22

    data_ref = Reference(ws, min_col=2, min_row=7, max_row=17)
    cats_ref = Reference(ws, min_col=1, min_row=8, max_row=17)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "D6")

    largura_auto(ws, {
        get_column_letter(i): 16 for i in range(1, 8)
    })
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 16


# ──────────────────────────────────────────────
# DASHBOARD HTML
# ──────────────────────────────────────────────
def gerar_dashboard_html(df: pd.DataFrame, caminho_saida: str, produtos_omie: dict = None):
    """Gera um dashboard HTML interativo com filtros de período, família e marca."""
    import json

    # ── Período para preencher os date inputs ───────────────────────
    dt_min = df["Data Emissão"].min()
    dt_max = df["Data Emissão"].max()
    dt_min_iso = dt_min.strftime("%Y-%m-%d")
    dt_max_iso = dt_max.strftime("%Y-%m-%d")
    periodo = f"{dt_min.strftime('%d/%m/%Y')} a {dt_max.strftime('%d/%m/%Y')}"

    # ── Listas para os selects de Família e Marca ───────────────────
    familias   = sorted([f for f in df["Família"].dropna().unique()
                        if f and f != "SEM CADASTRO"])
    marcas     = sorted([m for m in df["Marca"].dropna().unique()
                        if m and m != "SEM CADASTRO"])
    if "Vendedor" not in df.columns:
        df = df.copy()
        df["Vendedor"] = "Sem Vendedor"
    vendedores = sorted([v for v in df["Vendedor"].dropna().unique()
                        if v and v != "Sem Vendedor"])

    # ── Dados brutos para o JS (cada linha de venda) ────────────────
    cols_raw = ["NF", "Data Emissão", "Cliente", "Cód. Produto",
                "Produto", "Família", "Marca", "UF Dest.",
                "Qtd", "Vlr Bruto", "Desconto", "Vlr Líquido", "Vendedor"]
    raw = []
    for _, r in df[cols_raw].iterrows():
        raw.append({
            "nf":      str(r["NF"]),
            "data":    r["Data Emissão"].strftime("%Y-%m-%d"),
            "cliente": str(r["Cliente"])[:50],
            "cod":     str(r["Cód. Produto"]),
            "produto": str(r["Produto"])[:50],
            "familia": str(r["Família"]) if r["Família"] else "SEM CADASTRO",
            "marca":   str(r["Marca"])   if r["Marca"]   else "SEM CADASTRO",
            "uf":      str(r["UF Dest."]),
            "qtd":     round(float(r["Qtd"]), 4),
            "bruto":   round(float(r["Vlr Bruto"]), 2),
            "desc":    round(float(r["Desconto"]), 2),
            "liq":     round(float(r["Vlr Líquido"]), 2),
            "vendedor": str(r["Vendedor"]) if r["Vendedor"] else "Sem Vendedor",
        })


    def jv(v):
        return json.dumps(v, ensure_ascii=False)

    agora_str = datetime.now().strftime("%d/%m/%Y %H:%M")

    # ── Logo em base64 (embed no HTML, funciona sem dependência de arquivo) ──
    import base64, os as _os
    _base_dir = _os.path.dirname(_os.path.abspath(__file__))
    _logo_path = _os.path.join(_base_dir, "logo_purafor.jpg")
    _logo_b64 = ""
    if _os.path.exists(_logo_path):
        with open(_logo_path, "rb") as _f:
            _logo_b64 = base64.b64encode(_f.read()).decode()
    logo_tag = (f'<img src="data:image/jpeg;base64,{_logo_b64}" '
                f'style="height:56px;width:auto;border-radius:6px;'
                f'box-shadow:0 2px 8px rgba(0,0,0,.3);object-fit:contain;"/>'
                ) if _logo_b64 else ""

    # ── Logo Alfa Soluções em base64 ──
    _logo_alfa_path = _os.path.join(_base_dir, "Logo Alfa.jpg")
    _logo_alfa_b64 = ""
    if _os.path.exists(_logo_alfa_path):
        with open(_logo_alfa_path, "rb") as _f:
            _logo_alfa_b64 = base64.b64encode(_f.read()).decode()
    logo_alfa_tag = (f'<img src="data:image/jpeg;base64,{_logo_alfa_b64}" '
                    f'style="height:32px;width:auto;object-fit:contain;'
                    f'border-radius:4px;opacity:.85;"/>'
                    ) if _logo_alfa_b64 else "<strong>Alfa Soluções</strong>"

    cores_graficos = [
        "#2d3748","#4a5568","#059669","#d97706","#2563eb",
        "#7c3aed","#db2777","#0891b2","#65a30d","#dc2626",
        "#ea580c","#0d9488",
    ]

    # Opções HTML dos selects
    opt_fam  = "\n".join(f'<option value="{f}">{f}</option>' for f in familias)
    opt_marc = "\n".join(f'<option value="{m}">{m}</option>' for m in marcas)
    opt_vend = "\n".join(f'<option value="{v}">{v}</option>' for v in vendedores)

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0"/>
<title>Dashboard de Vendas — PURAFOR</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.2/dist/chart.umd.min.js"></script>
<style>
  *{{box-sizing:border-box;margin:0;padding:0;}}
  body{{font-family:'Segoe UI',Arial,sans-serif;background:#f8fafc;color:#1e293b;font-size:14px;}}

  /* ── TOPBAR ── */
  .topbar{{background:linear-gradient(135deg,#2d3748 0%,#1e293b 100%);
    color:#fff;padding:18px 32px;display:flex;align-items:center;justify-content:space-between;
    box-shadow:0 2px 8px rgba(0,0,0,.3);}}
  .topbar-title{{font-size:22px;font-weight:700;letter-spacing:1px;}}
  .topbar .sub{{font-size:12px;color:#a0aec0;margin-top:3px;}}
  .topbar .periodo{{font-size:12px;color:#90cdf4;text-align:right;}}

  /* ── FILTER BAR ── */
  .filter-bar{{background:#fff;border-bottom:2px solid #e2e8f0;
    padding:14px 32px;display:flex;align-items:flex-end;gap:16px;flex-wrap:wrap;
    box-shadow:0 1px 4px rgba(0,0,0,.07);position:sticky;top:0;z-index:100;}}
  .filter-group{{display:flex;flex-direction:column;gap:4px;}}
  .filter-group label{{font-size:11px;font-weight:700;color:#718096;text-transform:uppercase;
    letter-spacing:.4px;}}
  .filter-group input[type=date],
  .filter-group select{{border:1px solid #e2e8f0;border-radius:7px;padding:7px 12px;
    font-size:13px;color:#1e293b;background:#f8fafc;outline:none;
    min-width:150px;cursor:pointer;transition:border-color .15s;}}
  .filter-group input[type=date]:focus,
  .filter-group select:focus{{border-color:#2d3748;background:#fff;}}
  .filter-sep{{width:1px;height:38px;background:#e2e8f0;margin:0 4px;align-self:center;}}
  .btn{{padding:8px 20px;border:none;border-radius:7px;font-size:13px;font-weight:700;
    cursor:pointer;transition:all .15s;}}
  .btn-apply{{background:#2d3748;color:#fff;}}
  .btn-apply:hover{{background:#1e293b;}}
  .btn-clear{{background:#e2e8f0;color:#4a5568;}}
  .btn-clear:hover{{background:#cbd5e0;}}
  .filter-info{{font-size:12px;color:#059669;font-weight:600;margin-left:auto;
    align-self:center;white-space:nowrap;}}

  /* ── LAYOUT ── */
  .container{{max-width:1500px;margin:0 auto;padding:24px 20px;}}
  .section-title{{font-size:15px;font-weight:700;color:#2d3748;margin:28px 0 12px;
    padding-left:10px;border-left:4px solid #2d3748;letter-spacing:.5px;text-transform:uppercase;}}

  /* ── KPI CARDS ── */
  .kpi-grid{{display:grid;grid-template-columns:repeat(7,1fr);gap:14px;margin-bottom:28px;}}
  @media(max-width:1100px){{.kpi-grid{{grid-template-columns:repeat(4,1fr);}}}}
  @media(max-width:700px){{.kpi-grid{{grid-template-columns:repeat(2,1fr);}}}}
  .kpi-card{{background:#fff;border-radius:10px;padding:16px 14px;
    box-shadow:0 1px 4px rgba(0,0,0,.1);border-top:4px solid #2d3748;text-align:center;
    transition:transform .15s;}}
  .kpi-card:hover{{transform:translateY(-3px);box-shadow:0 4px 12px rgba(0,0,0,.15);}}
  .kpi-card.green{{border-top-color:#059669;}}
  .kpi-card.orange{{border-top-color:#d97706;}}
  .kpi-card.red{{border-top-color:#dc2626;}}
  .kpi-card.blue{{border-top-color:#2563eb;}}
  .kpi-label{{font-size:11px;color:#718096;font-weight:600;text-transform:uppercase;margin-bottom:6px;}}
  .kpi-value{{font-size:20px;font-weight:700;color:#1e293b;}}
  .kpi-value.small{{font-size:15px;}}

  /* ── CHARTS GRID ── */
  .chart-row{{display:grid;gap:18px;margin-bottom:18px;}}
  .chart-row.col2{{grid-template-columns:1fr 1fr;}}
  @media(max-width:900px){{.chart-row.col2{{grid-template-columns:1fr;}}}}
  .chart-card{{background:#fff;border-radius:10px;padding:18px 20px;
    box-shadow:0 1px 4px rgba(0,0,0,.1);}}
  .chart-card h3{{font-size:13px;font-weight:700;color:#4a5568;margin-bottom:14px;
    text-transform:uppercase;letter-spacing:.4px;}}
  .chart-wrap{{position:relative;}}

  /* ── CANAL TABLE ── */
  .canal-card{{background:#fff;border-radius:10px;padding:20px;
    box-shadow:0 1px 4px rgba(0,0,0,.1);margin-bottom:18px;}}
  .canal-card h3{{font-size:13px;font-weight:700;color:#4a5568;margin-bottom:14px;
    text-transform:uppercase;letter-spacing:.4px;}}
  .canal-wrap{{display:grid;grid-template-columns:1fr 1fr;gap:18px;}}
  @media(max-width:900px){{.canal-wrap{{grid-template-columns:1fr;}}}}
  /* filtro canal */
  .canal-filter{{display:flex;align-items:center;gap:10px;margin-bottom:16px;flex-wrap:wrap;}}
  .canal-filter span{{font-size:11px;font-weight:700;color:#718096;text-transform:uppercase;letter-spacing:.4px;margin-right:4px;}}
  .canal-toggle{{display:flex;align-items:center;gap:6px;padding:6px 14px;
    border-radius:20px;border:2px solid #e2e8f0;cursor:pointer;font-size:12px;
    font-weight:700;background:#fff;transition:all .15s;user-select:none;}}
  .canal-toggle input{{display:none;}}
  .canal-toggle.ativo-PURAFOR{{background:#2563eb;border-color:#2563eb;color:#fff;}}
  .canal-toggle.ativo-REAVITA{{background:#059669;border-color:#059669;color:#fff;}}
  .canal-toggle.ativo-TERCEIRIZADO{{background:#d97706;border-color:#d97706;color:#fff;}}
  .canal-toggle.ativo-OUTROS{{background:#6366f1;border-color:#6366f1;color:#fff;}}
  .canal-toggle:not([class*="ativo-"]){{color:#4a5568;}}
  #tblCanal{{width:100%;border-collapse:collapse;font-size:13px;}}
  #tblCanal th{{background:#3a7d44;color:#fff;padding:9px 14px;text-align:center;
    font-weight:700;font-size:12px;letter-spacing:.3px;white-space:nowrap;}}
  #tblCanal th.th-label{{background:#2d5a27;text-align:left;}}
  #tblCanal td{{padding:8px 14px;border-bottom:1px solid #e2e8f0;white-space:nowrap;}}
  #tblCanal td.num{{text-align:right;font-variant-numeric:tabular-nums;}}
  #tblCanal td.num.destaque{{font-weight:700;color:#2d5a27;}}
  #tblCanal td.num.azul{{font-weight:700;color:#2563eb;}}
  .grupo-label{{background:#d4edda;font-weight:800;color:#1a3d1f;
    padding:9px 14px;letter-spacing:.5px;text-transform:uppercase;font-size:12px;}}
  .sub-label{{padding-left:22px;color:#4a5568;}}
  .grupo-purafor td.grupo-label{{background:#c8e6c9;}}
  .grupo-reavita td.grupo-label{{background:#b3d9f7;color:#1a3a5c;}}
  .grupo-terc td.grupo-label{{background:#fff3cd;color:#5c4a00;}}

  /* ── RESUMO GRID ── */
  .resumo-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:28px;}}
  @media(max-width:900px){{.resumo-grid{{grid-template-columns:repeat(2,1fr);}}}}

  /* ── CANAL KPI CARDS ── */
  .canal-kpi-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:20px;margin-bottom:28px;}}
  @media(max-width:1200px){{.canal-kpi-grid{{grid-template-columns:repeat(2,1fr);}}}}
  @media(max-width:900px){{.canal-kpi-grid{{grid-template-columns:1fr;}}}}
  .canal-kpi-card{{background:#fff;border-radius:12px;padding:22px 20px;
    box-shadow:0 2px 8px rgba(0,0,0,.12);position:relative;overflow:hidden;}}
  .canal-kpi-card::before{{content:'';position:absolute;top:0;left:0;right:0;height:5px;}}
  .canal-kpi-card.purafor::before{{background:#2563eb;}}
  .canal-kpi-card.reavita::before{{background:#059669;}}
  .canal-kpi-card.terc::before{{background:#d97706;}}
  .canal-kpi-card.outros::before{{background:#6366f1;}}
  .canal-kpi-title{{font-size:10px;font-weight:800;text-transform:uppercase;letter-spacing:1.2px;
    padding:4px 12px;border-radius:20px;display:inline-block;margin-bottom:16px;}}
  .canal-kpi-card.purafor .canal-kpi-title{{background:#dbeafe;color:#1d4ed8;}}
  .canal-kpi-card.reavita .canal-kpi-title{{background:#d1fae5;color:#065f46;}}
  .canal-kpi-card.terc    .canal-kpi-title{{background:#fef3c7;color:#92400e;}}
  .canal-kpi-card.outros  .canal-kpi-title{{background:#ede9fe;color:#4c1d95;}}
  .canal-kpi-metrics{{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-top:4px;}}
  .ck-fat-row{{grid-column:span 2;}}
  .canal-kpi-metric .m-label{{font-size:10px;color:#718096;text-transform:uppercase;font-weight:600;margin-bottom:4px;}}
  .canal-kpi-metric .m-value{{font-size:20px;font-weight:800;color:#1e293b;line-height:1;}}
  .canal-kpi-metric .m-value.fat{{font-size:17px;}}
  .canal-kpi-metric .m-value.dev{{font-size:17px;color:#dc2626;}}
  .canal-kpi-share{{margin-top:16px;padding-top:14px;border-top:1px solid #f0f4f8;
    display:flex;align-items:center;gap:10px;}}
  .share-label{{font-size:11px;color:#718096;white-space:nowrap;font-weight:600;}}
  .share-bar-wrap{{flex:1;background:#f0f4f8;height:9px;border-radius:5px;overflow:hidden;}}
  .share-bar{{height:9px;border-radius:5px;width:0%;transition:width .5s ease;}}
  .share-pct{{font-size:15px;font-weight:800;white-space:nowrap;}}

  /* ── FOOTER ── */
  .footer{{
    margin-top:32px;padding:22px 32px;
    border-top:1px solid #e2e8f0;
    background:#f8fafc;
    display:flex;align-items:center;justify-content:space-between;
    gap:16px;flex-wrap:wrap;
  }}
  .footer-dev{{
    display:flex;align-items:center;gap:10px;
    font-size:12px;color:#94a3b8;line-height:1.5;
  }}
  .footer-dev strong{{color:#64748b;font-weight:600;}}
  .footer-sep{{width:1px;height:32px;background:#e2e8f0;}}
  .footer-gen{{font-size:11px;color:#b0bec5;text-align:right;}}

  /* ── BOTÃO VER PRODUTOS ── */
  .btn-ver-prod{{display:block;width:100%;margin-top:14px;padding:8px 0;
    border:none;border-radius:7px;font-size:12px;font-weight:700;cursor:pointer;
    letter-spacing:.4px;transition:all .15s;}}
  .purafor .btn-ver-prod{{background:#dbeafe;color:#1d4ed8;}}
  .purafor .btn-ver-prod:hover{{background:#2563eb;color:#fff;}}
  .reavita .btn-ver-prod{{background:#d1fae5;color:#065f46;}}
  .reavita .btn-ver-prod:hover{{background:#059669;color:#fff;}}
  .terc    .btn-ver-prod{{background:#fef3c7;color:#92400e;}}
  .terc    .btn-ver-prod:hover{{background:#d97706;color:#fff;}}
  .outros  .btn-ver-prod{{background:#ede9fe;color:#4c1d95;}}
  .outros  .btn-ver-prod:hover{{background:#6366f1;color:#fff;}}

  /* ── MODAL ── */
  .modal-overlay{{display:none;position:fixed;inset:0;background:rgba(0,0,0,.55);
    z-index:9000;justify-content:center;align-items:flex-start;padding:40px 16px;
    overflow-y:auto;}}
  .modal-overlay.aberto{{display:flex;}}
  .modal-box{{background:#fff;border-radius:14px;width:100%;max-width:820px;
    box-shadow:0 8px 40px rgba(0,0,0,.25);overflow:hidden;animation:slideIn .2s ease;}}
  @keyframes slideIn{{from{{transform:translateY(-30px);opacity:0}}to{{transform:translateY(0);opacity:1}}}}
  .modal-header{{padding:18px 24px;display:flex;align-items:center;justify-content:space-between;}}
  .modal-header.purafor{{background:#2563eb;}}
  .modal-header.reavita{{background:#059669;}}
  .modal-header.terc{{background:#d97706;}}
  .modal-header h2{{color:#fff;font-size:16px;font-weight:800;letter-spacing:.6px;}}
  .modal-close{{background:rgba(255,255,255,.25);border:none;color:#fff;font-size:20px;
    width:34px;height:34px;border-radius:50%;cursor:pointer;font-weight:700;
    display:flex;align-items:center;justify-content:center;transition:background .15s;}}
  .modal-close:hover{{background:rgba(255,255,255,.45);}}
  .modal-search{{padding:14px 20px;border-bottom:1px solid #e2e8f0;display:flex;gap:10px;align-items:center;}}
  .modal-search input{{flex:1;border:1px solid #e2e8f0;border-radius:7px;padding:8px 14px;
    font-size:13px;outline:none;transition:border-color .15s;}}
  .modal-search input:focus{{border-color:#64748b;}}
  .modal-body{{padding:0 20px 20px;max-height:60vh;overflow-y:auto;}}
  #modalTabela{{width:100%;border-collapse:collapse;font-size:13px;margin-top:4px;}}
  #modalTabela thead th{{position:sticky;top:0;background:#f8fafc;border-bottom:2px solid #e2e8f0;
    padding:10px 12px;text-align:left;font-size:11px;font-weight:700;color:#718096;
    text-transform:uppercase;letter-spacing:.4px;cursor:pointer;user-select:none;white-space:nowrap;}}
  #modalTabela thead th:hover{{background:#f1f5f9;}}
  #modalTabela thead th.sort-asc::after{{content:' ▲';font-size:9px;}}
  #modalTabela thead th.sort-desc::after{{content:' ▼';font-size:9px;}}
  #modalTabela tbody tr:hover{{background:#f8fafc;}}
  #modalTabela tbody td{{padding:9px 12px;border-bottom:1px solid #f0f4f8;
    vertical-align:middle;color:#1e293b;}}
  #modalTabela td.num{{text-align:right;font-variant-numeric:tabular-nums;font-weight:600;}}
  #modalTabela td.cod{{font-family:monospace;font-size:12px;color:#64748b;}}
  .modal-footer{{padding:12px 20px;border-top:1px solid #e2e8f0;font-size:12px;
    color:#718096;display:flex;justify-content:space-between;align-items:center;}}
  /* ── DARK MODE ── */
  #btn-theme{{position:fixed;bottom:24px;right:24px;z-index:10000;
    width:48px;height:48px;border-radius:50%;border:2px solid rgba(255,255,255,.18);
    cursor:pointer;font-size:21px;box-shadow:0 4px 18px rgba(0,0,0,.35);
    background:#334155;color:#f1f5f9;transition:all .2s;
    display:flex;align-items:center;justify-content:center;line-height:1;}}
  #btn-theme:hover{{transform:scale(1.12);box-shadow:0 6px 24px rgba(0,0,0,.5);}}
  body[data-theme="dark"]{{background:#0f172a;color:#e2e8f0;}}
  body[data-theme="dark"] .filter-bar{{background:#1e293b;border-color:#334155;box-shadow:0 1px 4px rgba(0,0,0,.4);}}
  body[data-theme="dark"] .filter-group label{{color:#94a3b8;}}
  body[data-theme="dark"] .filter-group input,
  body[data-theme="dark"] .filter-group select{{background:#0f172a;color:#e2e8f0;border-color:#475569;}}
  body[data-theme="dark"] .filter-group input:focus,
  body[data-theme="dark"] .filter-group select:focus{{background:#1e293b;border-color:#64748b;}}
  body[data-theme="dark"] .filter-sep{{background:#334155;}}
  body[data-theme="dark"] .btn-clear{{background:#334155;color:#cbd5e0;}}
  body[data-theme="dark"] .btn-clear:hover{{background:#475569;}}
  body[data-theme="dark"] .btn-apply{{background:#3b82f6;}}
  body[data-theme="dark"] .filter-info{{color:#34d399;}}
  body[data-theme="dark"] .section-title{{color:#94a3b8;border-color:#64748b;}}
  body[data-theme="dark"] .kpi-card{{background:#1e293b;box-shadow:0 2px 8px rgba(0,0,0,.5);}}
  body[data-theme="dark"] .kpi-label{{color:#94a3b8;}}
  body[data-theme="dark"] .kpi-value{{color:#f1f5f9;}}
  body[data-theme="dark"] .chart-card{{background:#1e293b;box-shadow:0 2px 8px rgba(0,0,0,.4);}}
  body[data-theme="dark"] .chart-card h3{{color:#94a3b8;}}
  body[data-theme="dark"] .canal-card{{background:#1e293b;box-shadow:0 2px 8px rgba(0,0,0,.4);}}
  body[data-theme="dark"] .canal-card h3{{color:#94a3b8;}}
  body[data-theme="dark"] .canal-kpi-card{{background:#1e293b;box-shadow:0 2px 12px rgba(0,0,0,.5);}}
  body[data-theme="dark"] .canal-kpi-metric .m-label{{color:#94a3b8;}}
  body[data-theme="dark"] .canal-kpi-metric .m-value{{color:#f1f5f9;}}
  body[data-theme="dark"] .canal-kpi-share{{border-color:#334155;}}
  body[data-theme="dark"] .share-bar-wrap{{background:#334155;}}
  body[data-theme="dark"] .share-label{{color:#94a3b8;}}
  body[data-theme="dark"] .dev-row{{border-color:#334155;}}
  body[data-theme="dark"] .canal-filter span{{color:#94a3b8;}}
  body[data-theme="dark"] .canal-toggle:not([class*="ativo-"]){{background:#1e293b;border-color:#475569;color:#94a3b8;}}
  body[data-theme="dark"] #tblCanal td{{border-color:#334155;color:#e2e8f0;}}
  body[data-theme="dark"] #tblCanal tbody tr:hover{{background:#334155;}}
  body[data-theme="dark"] .sub-label{{color:#94a3b8;}}
  body[data-theme="dark"] .grupo-label{{background:#1a2f1f;color:#86efac;}}
  body[data-theme="dark"] .grupo-purafor td.grupo-label{{background:#1e3566;color:#93c5fd;}}
  body[data-theme="dark"] .grupo-reavita td.grupo-label{{background:#1a3a2a;color:#6ee7b7;}}
  body[data-theme="dark"] .grupo-terc td.grupo-label{{background:#3a2800;color:#fcd34d;}}
  body[data-theme="dark"] .modal-box{{background:#1e293b;}}
  body[data-theme="dark"] .modal-search{{border-color:#334155;background:#1e293b;}}
  body[data-theme="dark"] .modal-search input{{background:#0f172a;color:#e2e8f0;border-color:#475569;}}
  body[data-theme="dark"] .modal-search input:focus{{border-color:#64748b;}}
  body[data-theme="dark"] .modal-body{{background:#1e293b;}}
  body[data-theme="dark"] #modalTabela thead th{{background:#0f172a;color:#94a3b8;border-color:#334155;}}
  body[data-theme="dark"] #modalTabela thead th:hover{{background:#1e293b;}}
  body[data-theme="dark"] #modalTabela tbody td{{color:#e2e8f0;border-color:#334155;}}
  body[data-theme="dark"] #modalTabela tbody tr:hover{{background:#334155;}}
  body[data-theme="dark"] .modal-footer{{border-color:#334155;color:#94a3b8;background:#1e293b;}}
  body[data-theme="dark"] .footer{{background:#1e293b;border-color:#334155;}}
  body[data-theme="dark"] .footer-dev{{color:#64748b;}}
  body[data-theme="dark"] .footer-dev strong{{color:#94a3b8;}}
  body[data-theme="dark"] .footer-gen{{color:#475569;}}
  body[data-theme="dark"] .footer-sep{{background:#334155;}}
  body[data-theme="dark"] #btn-theme{{background:#e2e8f0;color:#1e293b;border-color:rgba(0,0,0,.18);}}

</style>
</head>
<body>
<!-- BOTÃO DARK MODE -->
<button id="btn-theme" onclick="toggleTheme()" title="Alternar modo claro/escuro">🌙</button>

<!-- TOPBAR -->
<div class="topbar">
  <div style="display:flex;align-items:center;gap:18px;">
    {logo_tag}
    <div>
      <div class="topbar-title">📊 DASHBOARD DE VENDAS — PURAFOR</div>
      <div class="sub">Análise de desempenho por canal: PURAFOR · REAVITA · TERCEIRIZADO</div>
    </div>
  </div>
  <div class="periodo">
    Dados: <strong>{periodo}</strong><br/>
    <!-- Gerado em: {agora_str} -->
  </div>
</div>

<!-- BARRA DE FILTROS (sticky) -->
<div class="filter-bar">
  <div class="filter-group">
    <label>📅 Data Início</label>
    <input type="date" id="fDateIni" value="{dt_min_iso}"/>
  </div>
  <div class="filter-group">
    <label>📅 Data Fim</label>
    <input type="date" id="fDateFim" value="{dt_max_iso}"/>
  </div>
  <div class="filter-group">
    <label>Vendedor</label>
    <select id="fVend" onchange="aplicarFiltros()" style="min-width:160px">
      <option value="">Todos</option>
      {opt_vend}
    </select>
  </div>
  <div class="filter-sep"></div>
  <button class="btn btn-apply" onclick="aplicarFiltros()">▶ Aplicar</button>
  <button class="btn btn-clear" onclick="limparFiltros()">✕ Limpar</button>
  <div class="filter-info" id="filtroInfo"></div>
</div>

<div class="container">

  <!-- RESUMO GERAL -->
  <div class="section-title">Resumo — PURAFOR + REAVITA + TERCEIRIZADO</div>
  <div class="resumo-grid">
    <div class="kpi-card blue">
      <div class="kpi-label">Qtd. Total Vendida</div>
      <div class="kpi-value" id="kQtd">—</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">Clientes Ativos</div>
      <div class="kpi-value" id="kClientes">—</div>
    </div>
    <div class="kpi-card green">
      <div class="kpi-label">Fat. Líquido Total</div>
      <div class="kpi-value small" id="kLiq">—</div>
    </div>
    <div class="kpi-card orange">
      <div class="kpi-label">% Desconto Médio</div>
      <div class="kpi-value" id="kPerc">—</div>
    </div>
  </div>

  <!-- CANAL KPIs -->
  <div class="section-title">📊 Indicadores por Canal</div>
  <div class="canal-filter" style="margin-bottom:16px">
    <span>Exibir:</span>
    <label class="canal-toggle ativo-PURAFOR" id="lbl-PURAFOR" onclick="toggleCanal('PURAFOR',this)">
      <input type="checkbox" id="chk-PURAFOR" checked/> ● PURAFOR
    </label>
    <label class="canal-toggle ativo-REAVITA" id="lbl-REAVITA" onclick="toggleCanal('REAVITA',this)">
      <input type="checkbox" id="chk-REAVITA" checked/> ● REAVITA
    </label>
    <label class="canal-toggle ativo-TERCEIRIZADO" id="lbl-TERCEIRIZADO" onclick="toggleCanal('TERCEIRIZADO',this)">
      <input type="checkbox" id="chk-TERCEIRIZADO" checked/> ● TERCEIRIZADO
    </label>
    <label class="canal-toggle ativo-OUTROS" id="lbl-OUTROS" onclick="toggleCanal('OUTROS',this)">
      <input type="checkbox" id="chk-OUTROS" checked/> ● OUTROS
    </label>
  </div>
  <div class="canal-kpi-grid">

    <!-- PURAFOR -->
    <div class="canal-kpi-card purafor">
      <div class="canal-kpi-title">● PURAFOR</div>
      <div class="canal-kpi-metrics">
        <div class="canal-kpi-metric ck-fat-row">
          <div class="m-label">💰 Faturamento Líquido</div>
          <div class="m-value fat" id="ck-fat-PURAFOR">—</div>
        </div>
        <div class="canal-kpi-metric">
          <div class="m-label">📦 Quantidade</div>
          <div class="m-value" id="ck-qtd-PURAFOR">—</div>
        </div>
        <div class="canal-kpi-metric">
          <div class="m-label">💲 Preço Médio</div>
          <div class="m-value" id="ck-pm-PURAFOR">—</div>
        </div>
        <div class="canal-kpi-metric">
          <div class="m-label">🧾 NFs</div>
          <div class="m-value" id="ck-nfs-PURAFOR">—</div>
        </div>
        <div class="canal-kpi-metric">
          <div class="m-label">👥 Clientes</div>
          <div class="m-value" id="ck-cli-PURAFOR">—</div>
        </div>
      </div>
      <div class="canal-kpi-share">
        <span class="share-label">% do Total</span>
        <div class="share-bar-wrap"><div class="share-bar" id="ck-bar-PURAFOR"></div></div>
        <span class="share-pct" style="color:#2563eb" id="ck-share-PURAFOR">—</span>
      </div>
      <button class="btn-ver-prod" onclick="abrirModalProdutos('PURAFOR')">&#128269; Ver Produtos</button>
    </div>
    <div class="canal-kpi-card reavita">
      <div class="canal-kpi-title">● REAVITA</div>
      <div class="canal-kpi-metrics">
        <div class="canal-kpi-metric ck-fat-row">
          <div class="m-label">💰 Faturamento Líquido</div>
          <div class="m-value fat" id="ck-fat-REAVITA">—</div>
        </div>
        <div class="canal-kpi-metric">
          <div class="m-label">📦 Quantidade</div>
          <div class="m-value" id="ck-qtd-REAVITA">—</div>
        </div>
        <div class="canal-kpi-metric">
          <div class="m-label">💲 Preço Médio</div>
          <div class="m-value" id="ck-pm-REAVITA">—</div>
        </div>
        <div class="canal-kpi-metric">
          <div class="m-label">🧾 NFs</div>
          <div class="m-value" id="ck-nfs-REAVITA">—</div>
        </div>
        <div class="canal-kpi-metric">
          <div class="m-label">👥 Clientes</div>
          <div class="m-value" id="ck-cli-REAVITA">—</div>
        </div>
      </div>
      <div class="canal-kpi-share">
        <span class="share-label">% do Total</span>
        <div class="share-bar-wrap"><div class="share-bar" id="ck-bar-REAVITA"></div></div>
        <span class="share-pct" style="color:#059669" id="ck-share-REAVITA">—</span>
      </div>
      <button class="btn-ver-prod" onclick="abrirModalProdutos('REAVITA')">&#128269; Ver Produtos</button>
    </div>

    <!-- TERCEIRIZADO -->
    <div class="canal-kpi-card terc">
      <div class="canal-kpi-title">● TERCEIRIZADO</div>
      <div class="canal-kpi-metrics">
        <div class="canal-kpi-metric ck-fat-row">
          <div class="m-label">💰 Faturamento Líquido</div>
          <div class="m-value fat" id="ck-fat-TERCEIRIZADO">—</div>
        </div>
        <div class="canal-kpi-metric">
          <div class="m-label">📦 Quantidade</div>
          <div class="m-value" id="ck-qtd-TERCEIRIZADO">—</div>
        </div>
        <div class="canal-kpi-metric">
          <div class="m-label">💲 Preço Médio</div>
          <div class="m-value" id="ck-pm-TERCEIRIZADO">—</div>
        </div>
        <div class="canal-kpi-metric">
          <div class="m-label">🧾 NFs</div>
          <div class="m-value" id="ck-nfs-TERCEIRIZADO">—</div>
        </div>
        <div class="canal-kpi-metric">
          <div class="m-label">👥 Clientes</div>
          <div class="m-value" id="ck-cli-TERCEIRIZADO">—</div>
        </div>
      </div>
      <div class="canal-kpi-share">
        <span class="share-label">% do Total</span>
        <div class="share-bar-wrap"><div class="share-bar" id="ck-bar-TERCEIRIZADO"></div></div>
        <span class="share-pct" style="color:#d97706" id="ck-share-TERCEIRIZADO">—</span>
      </div>
      <button class="btn-ver-prod" onclick="abrirModalProdutos('TERCEIRIZADO')">&#128269; Ver Produtos</button>
    </div>

    <div class="canal-kpi-card outros">
      <div class="canal-kpi-title">● OUTROS</div>
      <div class="canal-kpi-metrics">
        <div class="canal-kpi-metric ck-fat-row">
          <div class="m-label">💰 Faturamento Líquido</div>
          <div class="m-value fat" id="ck-fat-OUTROS">—</div>
        </div>
        <div class="canal-kpi-metric">
          <div class="m-label">📦 Quantidade</div>
          <div class="m-value" id="ck-qtd-OUTROS">—</div>
        </div>
        <div class="canal-kpi-metric">
          <div class="m-label">💲 Preço Médio</div>
          <div class="m-value" id="ck-pm-OUTROS">—</div>
        </div>
        <div class="canal-kpi-metric">
          <div class="m-label">🧾 NFs</div>
          <div class="m-value" id="ck-nfs-OUTROS">—</div>
        </div>
        <div class="canal-kpi-metric">
          <div class="m-label">👥 Clientes</div>
          <div class="m-value" id="ck-cli-OUTROS">—</div>
        </div>
      </div>
      <div class="canal-kpi-share">
        <span class="share-label">% do Total</span>
        <div class="share-bar-wrap"><div class="share-bar" id="ck-bar-OUTROS"></div></div>
        <span class="share-pct" style="color:#6366f1" id="ck-share-OUTROS">—</span>
      </div>
      <button class="btn-ver-prod" onclick="abrirModalProdutos('OUTROS')">&#128269; Ver Produtos</button>
    </div>

  </div>

  <!-- PARTICIPAÇÃO -->
  <div class="section-title">Participação no Faturamento</div>
  <div class="chart-row col2" style="margin-bottom:18px">
    <div class="chart-card">
      <h3>🍩 Participação por Canal — Total do Período</h3>
      <div class="chart-wrap"><canvas id="chartParticipacao" height="130"></canvas></div>
    </div>
    <div class="chart-card">
      <h3>📊 Evolução da Participação % por Mês (Barras Empilhadas)</h3>
      <div class="chart-wrap"><canvas id="chartEvolucaoShare" height="130"></canvas></div>
    </div>
  </div>

  <!-- EVOLUÇÃO MENSAL -->
  <div class="section-title">Evolução Mensal</div>
  <div class="chart-card" style="margin-bottom:18px">
    <h3>📊 Faturamento por Canal — Barras Agrupadas por Mês</h3>
    <div class="chart-wrap"><canvas id="chartCanalBarra" height="70"></canvas></div>
  </div>
  <div class="canal-wrap" style="margin-bottom:18px">
    <div class="chart-card">
      <h3>💰 Faturamento por Canal — Linha Mensal</h3>
      <div class="chart-wrap"><canvas id="chartCanalFat" height="120"></canvas></div>
    </div>
    <div class="chart-card">
      <h3>📐 Preço Médio / Unidade por Canal</h3>
      <div class="chart-wrap"><canvas id="chartCanalPreco" height="120"></canvas></div>
    </div>
  </div>

  <!-- VENDAS POR VENDEDOR -->
  <div class="section-title">🧑‍💼 Vendas por Vendedor</div>
  <div style="display:flex;gap:20px;flex-wrap:wrap;margin-bottom:16px;">
    <div class="chart-wrap" style="flex:2 1 380px;"><canvas id="chartVendedor" height="200"></canvas></div>
    <div style="flex:1 1 280px;overflow-x:auto;">
      <table class="tbl-canal" style="width:100%;font-size:12px;">
        <thead><tr style="background:#f8fafc;">
          <th style="text-align:left;padding:6px 8px;">Vendedor</th>
          <th style="padding:6px 8px;">Fat. Líquido</th>
          <th style="padding:6px 8px;">Part.%</th>
          <th style="padding:6px 8px;">Qtd</th>
          <th style="padding:6px 8px;">Nº NFs</th>
        </tr></thead>
        <tbody id="tblVendedorBody"></tbody>
      </table>
    </div>
  </div>

  <!-- TABELA DETALHADA -->
  <div class="section-title">Detalhe por Canal / Mês</div>
  <div class="canal-card">
    <h3>📋 Faturamento, Quantidade e Preço Médio por Canal × Mês</h3>
    <div style="overflow-x:auto">
      <table id="tblCanal"><tbody id="tblCanalBody"></tbody></table>
    </div>
  </div>

</div>

<div class="footer">
  <div class="footer-dev">
    {logo_alfa_tag}
    <div>
      <div>Desenvolvido por <strong>Fabrício Zamprogno</strong></div>
      <div>em parceria com <strong>Alfa Soluções Consultoria</strong></div>
    </div>
  </div>
  <div class="footer-sep"></div>
  <div class="footer-gen">
    Dashboard PURAFOR<br/>
    <!-- Gerado em {agora_str} -->
  </div>
</div>

<!-- MODAL PRODUTOS -->
<div class="modal-overlay" id="modalOverlay" onclick="fecharModalSeFora(event)">
  <div class="modal-box">
    <div class="modal-header" id="modalHeader">
      <h2 id="modalTitulo">Produtos</h2>
      <button class="modal-close" onclick="fecharModal()">&#x2715;</button>
    </div>
    <div class="modal-search">
      <input type="text" id="modalBusca" placeholder="&#128269; Buscar por código ou descrição..." oninput="filtrarModal()"/>
    </div>
    <div class="modal-body">
      <table id="modalTabela">
        <thead>
          <tr>
            <th onclick="ordenarModal(0)" id="mth0">Código</th>
            <th onclick="ordenarModal(1)" id="mth1">Descrição do Produto</th>
            <th onclick="ordenarModal(2)" id="mth2" style="text-align:right">Qtd Vendida</th>
            <th onclick="ordenarModal(3)" id="mth3" style="text-align:right">Fat. Líquido</th>
            <th onclick="ordenarModal(4)" id="mth4" style="text-align:right">Preço Médio</th>
          </tr>
        </thead>
        <tbody id="modalCorpo"></tbody>
      </table>
    </div>
    <div class="modal-footer">
      <span id="modalInfo"></span>
      <span style="color:#94a3b8">Clique nos cabeçalhos para ordenar</span>
    </div>
  </div>
</div>

<script>
// ═══════════════════════════════════════════════════
//  DADOS BRUTOS (todas as linhas de venda)
// ═══════════════════════════════════════════════════
const DADOS     = {jv(raw)};
// Catálogo completo de produtos Omie (indexado pelo código da NF-e)
// Campos: codigo, descricao, descricao_familia, marca, ean, ncm, unidade,
//         valor_unitario, peso_bruto, peso_liq, inativo, tipoItem, imagens, etc.
const PRODUTOS_OMIE = {jv(produtos_omie or {{}})};
const CORES = {jv(cores_graficos)};
const BRL = v => 'R$\u00a0' + v.toLocaleString('pt-BR',{{minimumFractionDigits:2,maximumFractionDigits:2}});
const NUM = v => v.toLocaleString('pt-BR');


Chart.defaults.font.family = "'Segoe UI', Arial, sans-serif";
Chart.defaults.font.size   = 12;
Chart.defaults.color       = '#4a5568';

// ═══════════════════════════════════════════════════
//  ESTADO DOS GRÁFICOS (para destruir e recriar)
// ═══════════════════════════════════════════════════
const charts = {{}};

function destroyChart(id) {{
  if (charts[id]) {{ charts[id].destroy(); delete charts[id]; }}
}}

// ═══════════════════════════════════════════════════
//  AGREGAÇÕES DINÂMICAS
// ═══════════════════════════════════════════════════
function agrupar(rows, campo, top=10, excluir='SEM CADASTRO') {{
  const m = {{}};
  rows.forEach(r => {{
    const k = r[campo] || excluir;
    if (k === excluir) return;
    m[k] = (m[k]||0) + r.liq;
  }});
  return Object.entries(m)
    .sort((a,b)=>b[1]-a[1])
    .slice(0,top);
}}

function agruparMensal(rows) {{
  const m = {{}};
  rows.forEach(r => {{
    const ym = r.data.substring(0,7); // "YYYY-MM"
    m[ym] = (m[ym]||0) + r.liq;
  }});
  return Object.entries(m).sort((a,b)=>a[0]>b[0]?1:-1);
}}

// ═══════════════════════════════════════════════════
//  CANAL: PURAFOR / REAVITA / TERCEIRIZADO
// ═══════════════════════════════════════════════════
function canalDeRow(r) {{
  const m = (r.marca||'').toUpperCase();
  const f = (r.familia||'').toLowerCase();
  if (m === 'PURAFOR') return 'PURAFOR';
  if (m === 'REAVITA')          return 'REAVITA';
  if (f === 'terceirizado')     return 'TERCEIRIZADO';
  return 'OUTROS';
}}

function agruparCanal(rows) {{
  // Retorna {{canal: {{ym: {{liq, qtd}}}}}}
  const canais = ['TOTAL','PURAFOR','REAVITA','TERCEIRIZADO','OUTROS'];
  const dados = {{}};
  const meses = new Set();
  canais.forEach(c => dados[c] = {{}});
  rows.forEach(r => {{
    const ym = r.data.substring(0,7);
    meses.add(ym);
    // TOTAL
    if (!dados.TOTAL[ym]) dados.TOTAL[ym] = {{liq:0,qtd:0}};
    dados.TOTAL[ym].liq += r.liq;
    dados.TOTAL[ym].qtd += r.qtd;
    // Canal
    const c = canalDeRow(r);
    if (!dados[c]) dados[c] = {{}};
    if (!dados[c][ym]) dados[c][ym] = {{liq:0,qtd:0}};
    dados[c][ym].liq += r.liq;
    dados[c][ym].qtd += r.qtd;
  }});
  return {{dados, meses: [...meses].sort()}};
}}

const CANAL_CORES = {{
  'TOTAL':'#3a7d44','PURAFOR':'#2563eb','REAVITA':'#059669','TERCEIRIZADO':'#d97706','OUTROS':'#6366f1'
}};

function canaisAtivos() {{
  return ['PURAFOR','REAVITA','TERCEIRIZADO','OUTROS'].filter(c =>
    document.getElementById('chk-'+c) && document.getElementById('chk-'+c).checked
  );
}}

function toggleCanal(canal, lbl) {{
  const chk = document.getElementById('chk-'+canal);
  chk.checked = !chk.checked;
  if (chk.checked) {{
    lbl.classList.add('ativo-'+canal);
  }} else {{
    lbl.classList.remove('ativo-'+canal);
  }}
  atualizar();
}}

// ═// ═// ═// ═// ═// ═// ═// ═// ═// ═// ═// ═// ═// ═// ═// ═// ═// ═// ═// ═// ═// ═// ═// ═// ═
//  VENDAS POR VENDEDOR
// ═
//  VENDAS POR VENDEDOR
// ═
//  VENDAS POR VENDEDOR
// ═
//  VENDAS POR VENDEDOR
// ═
//  VENDAS POR VENDEDOR
// ═
//  VENDAS POR VENDEDOR
// ═
//  VENDAS POR VENDEDOR
// ═
//  VENDAS POR VENDEDOR
// ═
//  VENDAS POR VENDEDOR
// ═
//  VENDAS POR VENDEDOR
// ═
//  VENDAS POR VENDEDOR
// ═
//  VENDAS POR VENDEDOR
// ═
//  VENDAS POR VENDEDOR
// ═
//  VENDAS POR VENDEDOR
// ═
//  VENDAS POR VENDEDOR
// ═
//  VENDAS POR VENDEDOR
// ═
//  VENDAS POR VENDEDOR
// ═
//  VENDAS POR VENDEDOR
// ═
//  VENDAS POR VENDEDOR
// ═
//  VENDAS POR VENDEDOR
// ═
//  VENDAS POR VENDEDOR
// ═
//  VENDAS POR VENDEDOR
// ═
//  VENDAS POR VENDEDOR
// ═
//  VENDAS POR VENDEDOR
// ═
//  VENDAS POR VENDEDOR
// ═
function renderVendedor(rows) {{
  // Filtrar apenas canais ativos (igual ao atualizarKPIs)
  const ativos = canaisAtivos();
  rows = rows.filter(r => ativos.includes(canalDeRow(r)));
  const mapa = {{}};
  rows.forEach(r => {{
    const v = r.vendedor || 'Sem Vendedor';
    if (!mapa[v]) mapa[v] = {{liq:0, qtd:0, nfs: new Set()}};
    mapa[v].liq += r.liq;
    mapa[v].qtd += r.qtd;
    mapa[v].nfs.add(r.nf);
  }});
  const total = Object.values(mapa).reduce((s,x) => s + x.liq, 0);
  const sorted = Object.entries(mapa).sort((a,b) => b[1].liq - a[1].liq);

  let html = '';
  sorted.forEach(([nome, d]) => {{
    const part = total > 0 ? (d.liq / total * 100) : 0;
    html += `<tr>
      <td style='text-align:left;padding:5px 8px;'>${{nome}}</td>
      <td class='num destaque' style='padding:5px 8px;'>${{BRL(d.liq)}}</td>
      <td class='num' style='padding:5px 8px;'>${{part.toFixed(1)}}%</td>
      <td class='num azul' style='padding:5px 8px;'>${{NUM(Math.round(d.qtd))}}</td>
      <td class='num' style='padding:5px 8px;'>${{NUM(d.nfs.size)}}</td>
    </tr>`;
  }});
  document.getElementById('tblVendedorBody').innerHTML = html;

  destroyChart('chartVendedor');
  const canvas = document.getElementById('chartVendedor');
  if (!canvas) return;
  const labels = sorted.map(([n]) => n.length > 24 ? n.substring(0, 24) + '…' : n);
  const valores = sorted.map(([, d]) => Math.round(d.liq));
  const CORES_V = ['#2563eb','#059669','#d97706','#7c3aed','#db2777',
                   '#0891b2','#dc2626','#65a30d','#ea580c','#0d9488','#2d3748','#4a5568'];
  charts['chartVendedor'] = new Chart(canvas, {{
    type: 'bar',
    data: {{
      labels: labels,
      datasets: [{{
        label: 'Fat. Líquido',
        data: valores,
        backgroundColor: CORES_V.slice(0, sorted.length),
        borderRadius: 4,
        borderSkipped: false,
      }}]
    }},
    options: {{
      indexAxis: 'y',
      plugins: {{
        legend: {{ display: false }},
        tooltip: {{ callbacks: {{ label: c => BRL(c.raw) }} }}
      }},
      scales: {{
        x: {{
          ticks: {{ callback: v => 'R$' + (v >= 1000 ? (v / 1000).toFixed(0) + 'k' : v) }},
          grid: {{ color: '#e2e8f0' }}
        }},
        y: {{ grid: {{ display: false }} }}
      }}
    }}
  }});
}}

function renderCanalTable(rows) {{
  const ativos = canaisAtivos();
  const {{dados, meses}} = agruparCanal(rows);
  const mesesFmt = meses.map(ym => {{
    const [y,m] = ym.split('-');
    const nms = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez'];
    return nms[parseInt(m)-1]+'/'+y.substring(2);
  }});

  const estilos = {{'TOTAL':'grupo-total','PURAFOR':'grupo-purafor','REAVITA':'grupo-reavita','TERCEIRIZADO':'grupo-terc'}};

  let html = '<tr>';
  html += '<th class="th-label" colspan="2">Canal / Indicador</th>';
  mesesFmt.forEach(m => html += `<th>${{m}}</th>`);
  html += '</tr>';

  ['TOTAL',...ativos].forEach(canal => {{
    const est = estilos[canal];
    const d = dados[canal];
    html += `<tr class="${{est}}"><td class="grupo-label" colspan="${{2+meses.length}}">${{canal}}</td></tr>`;
    html += `<tr class="${{est}}"><td></td><td class="sub-label">Faturamento</td>`;
    meses.forEach(ym => {{
      const v = (d[ym]||{{liq:0}}).liq;
      html += `<td class="num destaque">${{BRL(v)}}</td>`;
    }});
    html += '</tr>';
    html += `<tr class="${{est}}"><td></td><td class="sub-label">Quantidade de produtos</td>`;
    meses.forEach(ym => {{
      const v = (d[ym]||{{qtd:0}}).qtd;
      html += `<td class="num azul">${{v.toLocaleString('pt-BR',{{maximumFractionDigits:0}})}}</td>`;
    }});
    html += '</tr>';
    html += `<tr class="${{est}}"><td></td><td class="sub-label">Preço Médio/Unidade</td>`;
    meses.forEach(ym => {{
      const dd = d[ym]||{{liq:0,qtd:0}};
      const pm = dd.qtd > 0 ? dd.liq/dd.qtd : 0;
      html += `<td class="num">${{BRL(pm)}}</td>`;
    }});
    html += '</tr>';
  }});

  document.getElementById('tblCanalBody').innerHTML = html;

  // Gráfico linha — Faturamento
  destroyChart('chartCanalFat');
  charts['chartCanalFat'] = new Chart(document.getElementById('chartCanalFat'), {{
    type:'line',
    data:{{
      labels: mesesFmt,
      datasets: ativos.map(c => ({{
        label: c,
        data: meses.map(ym => Math.round((dados[c][ym]||{{liq:0}}).liq)),
        borderColor: CANAL_CORES[c],
        backgroundColor: CANAL_CORES[c]+'22',
        tension:.3, fill:false,
        pointRadius:5, pointHoverRadius:7, borderWidth:2
      }}))
    }},
    options:{{
      plugins:{{legend:{{position:'bottom'}},
        tooltip:{{callbacks:{{label:c=>c.dataset.label+': '+BRL(c.raw)}}}}}},
      scales:{{
        y:{{ticks:{{callback:v=>'R$'+(v>=1000?(v/1000).toFixed(0)+'k':v)}},grid:{{color:'#e2e8f0'}}}},
        x:{{grid:{{color:'#e2e8f0'}}}}
      }}
    }}
  }});

  // Gráfico linha — Preço Médio
  destroyChart('chartCanalPreco');
  charts['chartCanalPreco'] = new Chart(document.getElementById('chartCanalPreco'), {{
    type:'line',
    data:{{
      labels: mesesFmt,
      datasets: ativos.map(c => ({{
        label: c,
        data: meses.map(ym => {{
          const dd = dados[c][ym]||{{liq:0,qtd:0}};
          return dd.qtd>0 ? Math.round((dd.liq/dd.qtd)*100)/100 : 0;
        }}),
        borderColor: CANAL_CORES[c],
        backgroundColor: CANAL_CORES[c]+'22',
        tension:.3, fill:false,
        pointRadius:5, pointHoverRadius:7, borderWidth:2
      }}))
    }},
    options:{{
      plugins:{{legend:{{position:'bottom'}},
        tooltip:{{callbacks:{{label:c=>c.dataset.label+': '+BRL(c.raw)}}}}}},
      scales:{{
        y:{{ticks:{{callback:v=>'R$ '+v.toFixed(2).replace('.',',')}},grid:{{color:'#e2e8f0'}}}},
        x:{{grid:{{color:'#e2e8f0'}}}}
      }}
    }}
  }});

  // Gráfico barras agrupadas por mês
  destroyChart('chartCanalBarra');
  charts['chartCanalBarra'] = new Chart(document.getElementById('chartCanalBarra'), {{
    type:'bar',
    data:{{
      labels: mesesFmt,
      datasets: ativos.map(c => ({{
        label: c,
        data: meses.map(ym => Math.round((dados[c][ym]||{{liq:0}}).liq)),
        backgroundColor: CANAL_CORES[c],
        borderRadius: 4, borderSkipped: false
      }}))
    }},
    options:{{
      plugins:{{legend:{{position:'bottom'}},
        tooltip:{{callbacks:{{label:c=>c.dataset.label+': '+BRL(c.raw)}}}}}},
      scales:{{
        y:{{ticks:{{callback:v=>'R$'+(v>=1000?(v/1000).toFixed(0)+'k':v)}},grid:{{color:'#e2e8f0'}}}},
        x:{{grid:{{display:false}}}}
      }}
    }}
  }});
}}

// ═══════════════════════════════════════════════════
//  MODAL DE PRODUTOS POR CANAL
// ═══════════════════════════════════════════════════
let _modalDados = [];  // dados agregados por produto do canal aberto
let _modalColSort = 3; // coluna de ordenação (padrão: Fat. Líquido)
let _modalSortAsc = false;

const CANAL_CLS = {{PURAFOR:'purafor',REAVITA:'reavita',TERCEIRIZADO:'terc'}};

function abrirModalProdutos(canal) {{
  // Agrega produtos do canal filtrado
  const rows  = dadosFiltrados.filter(r => canalDeRow(r) === canal);
  const mapa  = {{}};
  rows.forEach(r => {{
    const k = r.cod;
    if (!mapa[k]) mapa[k] = {{cod:r.cod, produto:r.produto, qtd:0, liq:0}};
    mapa[k].qtd += r.qtd;
    mapa[k].liq += r.liq;
  }});
  _modalDados = Object.values(mapa);
  _modalColSort = 3;
  _modalSortAsc = false;

  // Configura header do modal
  const cls = CANAL_CLS[canal] || '';
  const hdr = document.getElementById('modalHeader');
  hdr.className = 'modal-header ' + cls;
  document.getElementById('modalTitulo').textContent =
    '🔎 Produtos — ' + canal + ' (' + _modalDados.length + ' SKUs)';
  document.getElementById('modalBusca').value = '';

  renderizarModal();
  document.getElementById('modalOverlay').classList.add('aberto');
  document.body.style.overflow = 'hidden';
}}

function fecharModal() {{
  document.getElementById('modalOverlay').classList.remove('aberto');
  document.body.style.overflow = '';
}}

function fecharModalSeFora(e) {{
  if (e.target.id === 'modalOverlay') fecharModal();
}}

document.addEventListener('keydown', e => {{ if (e.key === 'Escape') fecharModal(); }});

function ordenarModal(col) {{
  if (_modalColSort === col) {{ _modalSortAsc = !_modalSortAsc; }}
  else {{ _modalColSort = col; _modalSortAsc = col < 2; }}
  renderizarModal();
}}

function filtrarModal() {{
  renderizarModal();
}}

function renderizarModal() {{
  const busca = (document.getElementById('modalBusca').value || '').toLowerCase();

  // Atualiza ícones de ordenação
  for (let i = 0; i <= 4; i++) {{
    const th = document.getElementById('mth'+i);
    if (th) {{
      th.classList.remove('sort-asc','sort-desc');
      if (i === _modalColSort) th.classList.add(_modalSortAsc ? 'sort-asc' : 'sort-desc');
    }}
  }}

  let dados = _modalDados.filter(d =>
    !busca ||
    d.cod.toLowerCase().includes(busca) ||
    d.produto.toLowerCase().includes(busca)
  );

  // Ordena
  dados.sort((a,b) => {{
    let va, vb;
    if (_modalColSort === 0)      {{ va = a.cod;     vb = b.cod; }}
    else if (_modalColSort === 1) {{ va = a.produto; vb = b.produto; }}
    else if (_modalColSort === 2) {{ va = a.qtd;     vb = b.qtd; }}
    else if (_modalColSort === 3) {{ va = a.liq;     vb = b.liq; }}
    else                          {{ va = a.qtd>0?a.liq/a.qtd:0; vb = b.qtd>0?b.liq/b.qtd:0; }}
    if (typeof va === 'string') return _modalSortAsc ? va.localeCompare(vb) : vb.localeCompare(va);
    return _modalSortAsc ? va-vb : vb-va;
  }});

  const totLiq = dados.reduce((s,d)=>s+d.liq,0);
  const totQtd = dados.reduce((s,d)=>s+d.qtd,0);

  let html = '';
  dados.forEach((d,i) => {{
    const pm = d.qtd > 0 ? d.liq/d.qtd : 0;
    const bg = i%2===0 ? '' : 'style="background:#fafafa"';
    html += `<tr ${{bg}}>`;
    html += `<td class="cod">${{d.cod}}</td>`;
    html += `<td>${{d.produto}}</td>`;
    html += `<td class="num">${{d.qtd.toLocaleString('pt-BR',{{maximumFractionDigits:0}})}}</td>`;
    html += `<td class="num">${{BRL(d.liq)}}</td>`;
    html += `<td class="num">${{BRL(pm)}}</td>`;
    html += '</tr>';
  }});

  // Rodapé de totais
  const pmTot = totQtd > 0 ? totLiq/totQtd : 0;
  html += `<tr style="background:#f0f4f8;font-weight:800;border-top:2px solid #e2e8f0">`;
  html += `<td colspan="2">TOTAL (${{dados.length}} produtos)</td>`;
  html += `<td class="num">${{totQtd.toLocaleString('pt-BR',{{maximumFractionDigits:0}})}}</td>`;
  html += `<td class="num">${{BRL(totLiq)}}</td>`;
  html += `<td class="num">${{BRL(pmTot)}}</td>`;
  html += '</tr>';

  document.getElementById('modalCorpo').innerHTML = html;
  document.getElementById('modalInfo').textContent =
    dados.length + ' de ' + _modalDados.length + ' produtos';
}}
let dadosFiltrados    = DADOS;

function filtrar() {{
  const ini = document.getElementById('fDateIni').value;
  const fim = document.getElementById('fDateFim').value;

  const vend = document.getElementById('fVend') ? document.getElementById('fVend').value : '';
  dadosFiltrados = DADOS.filter(r => {{
    if (ini && r.data < ini) return false;
    if (fim && r.data > fim) return false;
    if (vend && r.vendedor !== vend) return false;
    return true;
  }});


  const total = dadosFiltrados.length;
  document.getElementById('filtroInfo').textContent =
    total === DADOS.length ? '' :
    `✔ ${{NUM(total)}} de ${{NUM(DADOS.length)}} itens filtrados`;
}}

// ═══════════════════════════════════════════════════
//  KPIs GERAIS — só PURAFOR + REAVITA + TERCEIRIZADO
// ═══════════════════════════════════════════════════
function atualizarKPIs(rows) {{
  // KPIs gerais: usa TODOS os itens (inclusive sem canal cadastrado)
  // Os cards por canal (PURAFOR/REAVITA/TERCEIRIZADO) filtram individualmente
  const qtd      = rows.reduce((s,r)=>s+r.qtd,0);
  const clientes = new Set(rows.map(r=>r.cliente)).size;
  const bruto = rows.reduce((s,r)=>s+r.bruto,0);
  const desc  = rows.reduce((s,r)=>s+r.desc,0);
  const liq   = rows.reduce((s,r)=>s+r.liq,0);
  const perc  = bruto ? (desc/bruto*100) : 0;

  document.getElementById('kQtd').textContent     = NUM(Math.round(qtd));
  document.getElementById('kClientes').textContent = NUM(clientes);
  document.getElementById('kLiq').textContent      = BRL(liq);
  document.getElementById('kPerc').textContent     = perc.toFixed(1) + '%';
}}

// ═══════════════════════════════════════════════════
//  CANAL KPI CARDS
// ═══════════════════════════════════════════════════
function renderCanalKPIs(rows) {{
  const totLiq = rows.reduce((s,r)=>s+r.liq,0);
  ['PURAFOR','REAVITA','TERCEIRIZADO','OUTROS'].forEach(canal => {{
    const rC  = rows.filter(r => canalDeRow(r) === canal);
    const liq = rC.reduce((s,r)=>s+r.liq,0);
    const qtd = rC.reduce((s,r)=>s+r.qtd,0);
    const pm  = qtd > 0 ? liq/qtd : 0;
    const share = totLiq > 0 ? (liq/totLiq*100) : 0;
    const nfs = new Set(rC.map(r=>r.nf)).size;
    const cli = new Set(rC.map(r=>r.cliente)).size;


    const el = id => document.getElementById(id+'-'+canal);
    el('ck-fat').textContent    = BRL(liq);
    el('ck-qtd').textContent    = NUM(Math.round(qtd));
    el('ck-pm').textContent     = BRL(pm);
    el('ck-nfs').textContent    = NUM(nfs);
    el('ck-cli').textContent    = NUM(cli);
    el('ck-share').textContent  = share.toFixed(1) + '%';

    const cores = {{PURAFOR:'#2563eb',REAVITA:'#059669',TERCEIRIZADO:'#d97706'}};
    const bar = document.getElementById('ck-bar-'+canal);
    if (bar) {{ bar.style.width = Math.min(share,100).toFixed(1)+'%'; bar.style.background = cores[canal]; }}
  }});
}}

// ═══════════════════════════════════════════════════
//  PARTICIPAÇÃO POR CANAL — doughnut
// ═══════════════════════════════════════════════════
function mkParticipacao(rows) {{
  const canais = canaisAtivos();
  const vals   = canais.map(c => rows.filter(r=>canalDeRow(r)===c).reduce((s,r)=>s+r.liq,0));
  const tot    = vals.reduce((a,b)=>a+b,0);
  destroyChart('chartParticipacao');
  charts['chartParticipacao'] = new Chart(document.getElementById('chartParticipacao'), {{
    type:'doughnut',
    data:{{
      labels: canais,
      datasets:[{{
        data: vals,
        backgroundColor: canais.map(c=>CANAL_CORES[c]),
        borderWidth:3, borderColor:'#fff', hoverOffset:12
      }}]
    }},
    options:{{
      cutout:'58%',
      plugins:{{
        legend:{{position:'right',labels:{{font:{{size:13}},boxWidth:16,padding:16}}}},
        tooltip:{{callbacks:{{label:c=>c.label+': '+BRL(c.raw)+' ('+((tot?c.raw/tot*100:0).toFixed(1))+'%)'}}}}
      }}
    }}
  }});
}}

// ═══════════════════════════════════════════════════
//  EVOLUÇÃO SHARE % — stacked bar
// ═══════════════════════════════════════════════════
function mkShareEvolucao(rows) {{
  const {{dados,meses}} = agruparCanal(rows);
  const canais = canaisAtivos();
  const mesesFmt = meses.map(ym => {{
    const [y,m] = ym.split('-');
    const ns = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez'];
    return ns[parseInt(m)-1]+'/'+y.substring(2);
  }});
  destroyChart('chartEvolucaoShare');
  charts['chartEvolucaoShare'] = new Chart(document.getElementById('chartEvolucaoShare'), {{
    type:'bar',
    data:{{
      labels: mesesFmt,
      datasets: canais.map(c => ({{
        label: c,
        data: meses.map(ym => {{
          // total do mês = soma apenas dos canais ativos
          const totMes = canais.reduce((s,cc)=>s+((dados[cc][ym]||{{liq:0}}).liq),0);
          const val    = (dados[c][ym]||{{liq:0}}).liq;
          return totMes > 0 ? Math.round(val/totMes*1000)/10 : 0;
        }}),
        backgroundColor: CANAL_CORES[c],
        borderRadius:2, borderSkipped:false
      }}))
    }},
    options:{{
      plugins:{{
        legend:{{position:'bottom'}},
        tooltip:{{callbacks:{{label:c=>c.dataset.label+': '+c.raw.toFixed(1)+'%'}}}}
      }},
      scales:{{
        x:{{stacked:true,grid:{{display:false}}}},
        y:{{stacked:true,max:100,ticks:{{callback:v=>v+'%'}},grid:{{color:'#e2e8f0'}}}}
      }}
    }}
  }});
}}

// ═══════════════════════════════════════════════════
//  GRÁFICOS
// ═══════════════════════════════════════════════════
function mkHorizBar(id, entries, color) {{
  destroyChart(id);
  const labels = entries.map(e=>e[0]);
  const data   = entries.map(e=>e[1]);
  charts[id] = new Chart(document.getElementById(id), {{
    type:'bar',
    data:{{ labels, datasets:[{{ data, backgroundColor: color||CORES,
      borderRadius:5, borderSkipped:false }}] }},
    options:{{
      indexAxis:'y',
      plugins:{{ legend:{{display:false}},
        tooltip:{{callbacks:{{label:c=>BRL(c.raw)}}}} }},
      scales:{{
        x:{{ticks:{{callback:v=>'R$'+(v>=1000?(v/1000).toFixed(0)+'k':v.toFixed(0))}},grid:{{color:'#e2e8f0'}}}},
        y:{{grid:{{display:false}}}}
      }}
    }}
  }});
}}

function mkVertBar(id, entries, color) {{
  destroyChart(id);
  const labels = entries.map(e=>e[0]);
  const data   = entries.map(e=>e[1]);
  charts[id] = new Chart(document.getElementById(id), {{
    type:'bar',
    data:{{ labels, datasets:[{{ data, backgroundColor:color||CORES,
      borderRadius:4, borderSkipped:false }}] }},
    options:{{
      plugins:{{ legend:{{display:false}},
        tooltip:{{callbacks:{{label:c=>BRL(c.raw)}}}} }},
      scales:{{
        y:{{ticks:{{callback:v=>'R$'+(v>=1000?(v/1000).toFixed(0)+'k':v.toFixed(0))}},grid:{{color:'#e2e8f0'}}}},
        x:{{grid:{{display:false}},ticks:{{maxRotation:40}}}}
      }}
    }}
  }});
}}

function mkLinha(id, entries) {{
  destroyChart(id);
  // Formata labels como "Jan/2024"
  const meses = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez'];
  const labels = entries.map(e=>{{ const p=e[0].split('-'); return meses[parseInt(p[1])-1]+'/'+p[0]; }});
  const data   = entries.map(e=>e[1]);
  charts[id] = new Chart(document.getElementById(id), {{
    type:'line',
    data:{{ labels, datasets:[{{
      label:'Fat. Líquido', data,
      borderColor:'#059669', backgroundColor:'rgba(5,150,105,.12)',
      tension:.3, fill:true,
      pointBackgroundColor:'#059669', pointRadius:5, pointHoverRadius:7
    }}] }},
    options:{{
      plugins:{{ legend:{{display:false}},
        tooltip:{{callbacks:{{label:c=>BRL(c.raw)}}}} }},
      scales:{{
        y:{{ticks:{{callback:v=>'R$'+(v>=1000?(v/1000).toFixed(0)+'k':v.toFixed(0))}},grid:{{color:'#e2e8f0'}}}},
        x:{{grid:{{color:'#e2e8f0'}}}}
      }}
    }}
  }});
}}

function mkDoughnut(id, entries) {{
  destroyChart(id);
  charts[id] = new Chart(document.getElementById(id), {{
    type:'doughnut',
    data:{{ labels:entries.map(e=>e[0]),
      datasets:[{{ data:entries.map(e=>e[1]), backgroundColor:CORES,
        borderWidth:2, borderColor:'#fff', hoverOffset:8 }}] }},
    options:{{
      cutout:'55%',
      plugins:{{
        legend:{{position:'right',labels:{{font:{{size:11}},boxWidth:14}}}},
        tooltip:{{callbacks:{{label:c=>c.label+': '+BRL(c.raw)}}}}
      }}
    }}
  }});
}}

// ═══════════════════════════════════════════════════
//  ATUALIZAÇÃO GERAL
// ═══════════════════════════════════════════════════
function atualizar() {{
  const rows = dadosFiltrados;
  atualizarKPIs(rows);
  renderCanalKPIs(rows);
  mkParticipacao(rows);
  mkShareEvolucao(rows);
  renderCanalTable(rows);
  renderVendedor(rows);
}}

function aplicarFiltros() {{ filtrar(); atualizar(); }}

function limparFiltros() {{
  document.getElementById('fDateIni').value = '{dt_min_iso}';
  document.getElementById('fDateFim').value = '{dt_max_iso}';
  document.getElementById('filtroInfo').textContent = '';
  if (document.getElementById('fVend')) document.getElementById('fVend').value = '';
  dadosFiltrados    = DADOS;
  atualizar();
}}

// ── Inicializa ──────────────────────────────────────
dadosFiltrados    = DADOS;
atualizar();
// ── DARK MODE ─────────────────────────────────────
function toggleTheme() {{
  const body = document.body;
  const isDark = body.getAttribute('data-theme') === 'dark';
  const next = isDark ? 'light' : 'dark';
  body.setAttribute('data-theme', next);
  document.getElementById('btn-theme').textContent = isDark ? '\u{1F319}' : '\u2600\uFE0F';
  try {{ localStorage.setItem('purafor-theme', next); }} catch(e) {{}}
  const gridColor  = next === 'dark' ? '#334155' : '#e2e8f0';
  const labelColor = next === 'dark' ? '#94a3b8' : '#4a5568';
  Chart.defaults.color = labelColor;
  Chart.defaults.borderColor = gridColor;
  Object.values(Chart.instances).forEach(ch => {{
    if (ch.options.scales) {{
      Object.values(ch.options.scales).forEach(sc => {{
        if (sc.grid)  sc.grid.color  = gridColor;
        if (sc.ticks) sc.ticks.color = labelColor;
      }});
    }}
    if (ch.options.plugins && ch.options.plugins.legend)
      ch.options.plugins.legend.labels.color = labelColor;
    ch.update();
  }});
}}

document.addEventListener('DOMContentLoaded', () => {{
  const pg = document.getElementById('pg-inp');
  if (pg) pg.focus();
  try {{
    const saved = localStorage.getItem('purafor-theme');
    if (saved === 'dark') {{
      document.body.setAttribute('data-theme', 'dark');
      const btn = document.getElementById('btn-theme');
      if (btn) btn.textContent = '\u2600\uFE0F';
    }}
  }} catch(e) {{}}
}});

</script>
</body>
</html>"""

    with open(caminho_saida, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  ✔ Dashboard HTML salvo em:\n    {caminho_saida}")


# ──────────────────────────────────────────────
# MAPA VENDEDOR (Omie: ListarPedidos + ListarVendedores)
# ──────────────────────────────────────────────
def _buscar_mapa_vendedor(data_ini: str, data_fim: str) -> dict:
    """
    Retorna {cChaveNFe (str, 44 digitos): nome_vendedor (str)}.

    Estrategia direta via NFConsultar (Vendas e NF-e):
      1. ListarVendedores  -> {codigo -> nome}
      2. ListarNF (cDetalhesPedido="S") -> {cChaveNFe -> nome_vendedor}
         Usa pedido.nIdVendedor -- campo nativo da NF, sem precisar de CR.
         Fallback: titulos[0].nCodVendedor se nIdVendedor ausente.

    Cache de modulo: 3 horas.
    """
    import time as _vt

    # Garante que o global existe mesmo apos importlib.reload parcial
    global _VENDOR_MAP_CACHE
    if '_VENDOR_MAP_CACHE' not in globals() or not isinstance(_VENDOR_MAP_CACHE, dict):
        _VENDOR_MAP_CACHE = {}

    URL_VEND = 'https://app.omie.com.br/api/v1/geral/vendedores/'
    URL_NF   = 'https://app.omie.com.br/api/v1/produtos/nfconsultar/'

    # Cache de modulo (3 horas)
    _cache_key = f"{data_ini}|{data_fim}"
    _cached = _VENDOR_MAP_CACHE.get(_cache_key)
    if _cached is not None:
        ts, resultado = _cached
        if _vt.time() - ts < 10800:
            print(f"  \u2714 Mapa de vendedores do cache ({len(resultado)} entradas)")
            return resultado

    # 1. Busca todos os vendedores: {codigo -> nome}
    mapa_vend: dict = {}
    pag = 1
    while True:
        try:
            r = requests.post(URL_VEND, json={
                'call': 'ListarVendedores', 'app_key': OMIE_APP_KEY,
                'app_secret': OMIE_APP_SECRET,
                'param': [{'pagina': pag, 'registros_por_pagina': 500}]  # máximo Omie
            }, timeout=30).json()
            for v in r.get('cadastro', []):
                mapa_vend[v['codigo']] = v['nome']
            if pag >= r.get('total_de_paginas', 1):
                break
            pag += 1
        except Exception as e:
            print(f"  [AVISO] Erro ao buscar vendedores (pag {pag}): {e}")
            break
    print(f"  \u2714 {len(mapa_vend)} vendedores cadastrados na Omie")

    # 2. ListarNF filtrado por data de emissao: {cChaveNFe -> nome_vendedor}
    #    pedido.nIdVendedor e campo nativo da NF -- direto, sem chaining por CR.
    mapa_chave_vend: dict = {}
    pag = 1
    while True:
        _vt.sleep(0.3)  # rate limit Omie (ListarNF nao tem lock como CR)
        try:
            r = requests.post(URL_NF, json={
                'call': 'ListarNF', 'app_key': OMIE_APP_KEY,
                'app_secret': OMIE_APP_SECRET,
                'param': [{
                    'pagina':              pag,
                    'registros_por_pagina': 100,
                    'dEmiInicial':         data_ini,
                    'dEmiFinal':           data_fim,
                    'tpNF':                '1',   # apenas saida (vendas)
                    'cDetalhesPedido':     'S',   # inclui pedido.nIdVendedor
                }]
            }, timeout=60).json()

            if 'faultstring' in r or 'faultcode' in r:
                print(f"  [AVISO] Omie erro ListarNF pag {pag}: {r.get('faultstring', r)}")
                break

            nfs = r.get('nfCadastro', [])
            for nf in nfs:
                chave    = (nf.get('compl') or {}).get('cChaveNFe', '')
                nid_vend = (nf.get('pedido') or {}).get('nIdVendedor', 0)
                # fallback: primeiro titulo da NF
                if not nid_vend:
                    titulos = nf.get('titulos', []) or []
                    if titulos:
                        nid_vend = titulos[0].get('nCodVendedor', 0)
                if chave and nid_vend:
                    mapa_chave_vend[chave] = mapa_vend.get(
                        int(nid_vend), f'Vendedor-{nid_vend}')

            total_pag = r.get('total_de_paginas', 1)
            print(f"  NF pag {pag}/{total_pag}: {len(nfs)} NFs processadas")
            if pag >= total_pag:
                break
            pag += 1
        except Exception as e:
            print(f"  [AVISO] Erro ao buscar NFs (pag {pag}): {e}")
            break

    print(f"  \u2714 {len(mapa_chave_vend)} NFs com vendedor identificado via ListarNF")
    _VENDOR_MAP_CACHE[_cache_key] = (_vt.time(), mapa_chave_vend)
    return mapa_chave_vend


# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────
_EXCEL_DEFAULT = object()  # sentinel

def main(
    saida_html:  str | None = None,
    saida_excel = _EXCEL_DEFAULT,
    data_ini:    str | None = None,
    data_fim:    str | None = None,
) -> str | None:
    """
    Executa coleta + geração do dashboard.
    - saida_html:  caminho do HTML a gerar (usa SAIDA_HTML global se omitido)
    - saida_excel: None = pula Excel; não fornecido = usa SAIDA_EXCEL global
    - data_ini / data_fim: período no formato 'DD/MM/AAAA' (usa globals se omitido)
    - Retorna o conteúdo HTML como string (útil para exibir no Streamlit).
    """
    _html_path  = saida_html or SAIDA_HTML
    _excel_path = SAIDA_EXCEL if saida_excel is _EXCEL_DEFAULT else saida_excel
    _data_ini   = data_ini or OMIE_DATA_INI
    _data_fim   = data_fim or OMIE_DATA_FIM
    print("=" * 55)
    print("  RELATÓRIO DE VENDAS — PURAFOR")
    print("=" * 55)
    print(f"\nBuscando NF-e na API Omie: {_data_ini} a {_data_fim}")
    _prog(0.02, f"Iniciando coleta: {_data_ini} → {_data_fim}...")

    registros = _ler_vendas_com_cache(_data_ini, _data_fim)
    if not registros:
        print("\n[ERRO] Nenhum registro de venda encontrado!")
        return None

    df = pd.DataFrame(registros)
    df["Data Emissão"] = pd.to_datetime(df["Data Emissão"])

    print(f"  ✔ {len(registros)} itens de venda em {df['NF'].nunique()} NFs")
    print(f"  ✔ {df['Cliente'].nunique()} clientes distintos")
    print(f"  ✔ {df['Cód. Produto'].nunique()} produtos distintos")
    print(f"  ✔ Faturamento Líquido Total: R$ {df['Vlr Líquido'].sum():,.2f}")

    # ── JOIN com catálogo de Família/Marca ──────────────────────────
    _prog(0.40, "Carregando catálogo de produtos...")
    print("\nCarregando catálogo de produtos (Omie API)...")
    omie_map = carregar_catalogo_omie()

    print("Carregando catálogo de produtos (Excel fallback)...")
    df_cat = carregar_catalogo(CATALOGO_XLSX)
    mapa_excel = {}
    if not df_cat.empty:
        mapa_excel = df_cat.set_index("Codigo")[["Familia", "Marca"]].to_dict("index")

    # Dict cod_xml → produto Omie completo, usado para embutir no HTML
    produto_omie_por_xml: dict = {}

    def enriquecer(row):
        cod = str(row["Cód. Produto"]).strip()
        cod_n = _norm_cod(cod)

        # 1) Tenta no Omie (já inclui variações sem prefixo UN/CX)
        p_omie = omie_map.get(cod_n)
        fam, marc = "", ""
        if p_omie:
            produto_omie_por_xml[cod] = p_omie
            fam  = str(p_omie.get('descricao_familia', '') or '').strip()
            marc = str(p_omie.get('marca', '') or '').strip()

        # 2) Se Omie não tem familia/marca, tenta catálogo Excel
        if not fam or not marc:
            info = mapa_excel.get(cod, {})
            fam  = fam  or str(info.get("Familia", "") or "").strip()
            marc = marc or str(info.get("Marca",   "") or "").strip()

        row["Família"] = fam  or "SEM CADASTRO"
        row["Marca"]   = marc or "SEM CADASTRO"
        return row

    df = df.apply(enriquecer, axis=1)

    com_familia = (df["Família"] != "SEM CADASTRO").sum()
    print(f"  ✔ {com_familia:,} itens COM Família/Marca ({com_familia/len(df)*100:.1f}%)")

    # ── JOIN com mapa de Vendedores ────────────────────────────
    _prog(0.44, 'Buscando mapa de vendedores...')
    print("\nBuscando mapa de vendedores (Omie API)...")
    try:
        mapa_vendedor = _buscar_mapa_vendedor(_data_ini, _data_fim)
        # 'nChave' pode estar ausente em cache antigo — preenche com '' se faltar
        if 'nChave' not in df.columns:
            df['nChave'] = ''
        # Garante nChave como string limpa (NaN vira '')
        df['nChave'] = df['nChave'].fillna('').astype(str)
        # Aplica o mapa ANTES de qualquer diagnostico (evita bug no except)
        df['Vendedor'] = df['nChave'].map(mapa_vendedor).fillna('Sem Vendedor')
        # Diagnostico
        n_com_chave = int((df['nChave'] != '').sum())
        n_vend = int((df['Vendedor'] != 'Sem Vendedor').sum())
        print(f"  Diagnostico JOIN: {len(df)} itens | {n_com_chave} com nChave | "
              f"{len(mapa_vendedor)} entradas no mapa | {n_vend} com vendedor ")
        sample = df.loc[df['nChave'] != '', 'nChave'].head(3).tolist()
        for ch in sample:
            res = mapa_vendedor.get(str(ch), 'NAO ENCONTRADO')
            print(f"    nChave={str(ch)[:22]}... -> {res}")
        print(f"  {df['Vendedor'].nunique()} vendedores | "
              f"{n_vend}/{len(df)} ({100 * n_vend / max(len(df), 1):.1f}%)")
    except Exception as _e_vend:
        import traceback; traceback.print_exc()
        print(f'  [AVISO] Erro ao buscar vendedores: {_e_vend}')
        df['Vendedor'] = 'Sem Vendedor'

    print("\nGerando planilha Excel...")

    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove sheet padrão

    sheet_detalhe(wb, df)
    grp_produto = sheet_por_produto(wb, df)
    sheet_por_familia(wb, df)
    sheet_por_marca(wb, df)
    sheet_por_cliente(wb, df)
    sheet_por_data(wb, df)
    sheet_depara(wb, df)
    sheet_dashboard(wb, df, grp_produto)

    # Reordena abas
    order = ["Dashboard", "Vendas por Família", "Vendas por Marca",
             "Vendas por Produto", "Vendas por Cliente",
             "Vendas por Data", "De-Para (Preencher)", "Detalhe de Vendas"]
    for i, name in enumerate(order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=wb.sheetnames.index(name) - i)

    if _excel_path:
        wb.save(_excel_path)
        print(f"\n  ✔ Excel salvo em:\n    {_excel_path}")
    else:
        print("\n  (Excel: não gerar, modo cloud)")


    _prog(0.88, "Gerando Dashboard HTML...")
    print("\nGerando Dashboard HTML...")
    gerar_dashboard_html(df, _html_path, produtos_omie=produto_omie_por_xml)

    _prog(1.0, "Concluído!")
    print("\nPronto!")

    # Retorna o conteúdo HTML para uso pelo Streamlit
    try:
        with open(_html_path, encoding='utf-8') as _f:
            return _f.read()
    except Exception:
        return None


if __name__ == "__main__":
    main()

